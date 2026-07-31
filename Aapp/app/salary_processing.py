"""
Salary Processing Module
========================
Handles monthly salary processing, payslip generation, bulk operations and
statutory compliance (PF/ESI/PT/IT).

Pay scale source of truth: Designation (Aapp/app/designation.py). Each
employee is assigned a designation, and that designation carries the full
pay structure (basic, HRA, DA, allowances) plus statutory rates (PF/ESI/PT/IT,
both employee and employer share). There is no separate per-employee salary
structure — changing pay means editing the designation (or moving the
employee to a different designation), matching how designation.py is built.
"""

import calendar
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime

from django.contrib import messages
from django.db import transaction
from django.db import models
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from Aapp.app.employee import employee
from Aapp.app.designation import designation
from Sapp.app.company import Company as company

MONTH_CHOICES = [
    (1,'January'),(2,'February'),(3,'March'),(4,'April'),
    (5,'May'),(6,'June'),(7,'July'),(8,'August'),
    (9,'September'),(10,'October'),(11,'November'),(12,'December'),
]
YEAR_CHOICES = [(y) for y in range(2026, 2032)]

#======================================================================
#       Salary model tables
#======================================================================


class salary_processing(models.Model):
    """Monthly salary processing batch."""
    company_id = models.ForeignKey(company, on_delete=models.CASCADE, db_column="CompanyID")
    month = models.IntegerField()  # 1-12
    year = models.IntegerField()

    total_employees = models.IntegerField(default=0)
    total_gross = models.DecimalField(max_digits=15, decimal_places=2, default=0)
    total_deductions = models.DecimalField(max_digits=15, decimal_places=2, default=0)
    total_net = models.DecimalField(max_digits=15, decimal_places=2, default=0)
    total_employer_cost = models.DecimalField(max_digits=15, decimal_places=2, default=0)

    status = models.CharField(max_length=20, default='DRAFT')  # DRAFT, PROCESSED, APPROVED, PAID

    processed_at = models.DateTimeField(null=True, blank=True)
    approved_at = models.DateTimeField(null=True, blank=True)

    created_at = models.DateTimeField(auto_now_add=True)
    created_by = models.CharField(max_length=50, null=True, blank=True)
    updated_at = models.DateTimeField(auto_now=True)
    updated_by = models.CharField(max_length=50, null=True, blank=True)

    class Meta:
        db_table = 'salary_processing'
        unique_together = ['company_id', 'month', 'year']


class salary_slip(models.Model):
    """Individual employee salary slip for a given batch."""
    processing_id = models.ForeignKey(salary_processing, on_delete=models.CASCADE, db_column="ProcessingID")
    employee_id = models.ForeignKey(employee, on_delete=models.CASCADE, db_column="EmployeeID")
    company_id = models.ForeignKey(company, on_delete=models.CASCADE, db_column="CompanyID")
    designation_id = models.ForeignKey(designation, on_delete=models.SET_NULL, null=True, db_column="DesignationID")

    # Attendance (mirrors Aapp.app.attandance.attendance for this month)
    total_days = models.IntegerField(default=0)
    working_days = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    paid_days = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    leave_days = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    overtime_hours = models.DecimalField(max_digits=6, decimal_places=2, default=0)

    # Earnings (Calculated, sourced from designation pay scale)
    basic_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    hra_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    da_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    conveyance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    medical_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    lunch_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    cca_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    special_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    travel_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    washing_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    cycle_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    other_allowance_earned = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    overtime_amount = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    bonus_amount = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    gross_earnings = models.DecimalField(max_digits=10, decimal_places=2, default=0)

    # Deductions (Calculated, sourced from designation statutory rates)
    pf_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    esi_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    labour_welfare_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    professional_tax = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    income_tax = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    loan_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    advance_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    late_fine = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    other_deduction = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    total_deductions = models.DecimalField(max_digits=10, decimal_places=2, default=0)

    net_pay = models.DecimalField(max_digits=10, decimal_places=2, default=0)

    # Employer Cost
    pf_employer_contribution = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    esi_employer_contribution = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    labour_welfare_employer_contribution = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    employer_total_cost = models.DecimalField(max_digits=10, decimal_places=2, default=0)

    created_at = models.DateTimeField(auto_now_add=True)
    created_by = models.CharField(max_length=50, null=True, blank=True)
    updated_at = models.DateTimeField(auto_now=True)
    updated_by = models.CharField(max_length=50, null=True, blank=True)

    class Meta:
        db_table = 'salary_slip'
        unique_together = ['processing_id', 'employee_id']


# =====================================================================
# CONSTANTS
# =====================================================================

PF_WAGE_CEILING = Decimal('15000')       # Max basic for PF calculation
ESI_WAGE_CEILING = Decimal('21000')      # Max gross for ESI applicability
STANDARD_WORKING_HOURS = Decimal('8')    # Hours per day, for OT rate derivation

PT_STATES = {                            # Professional Tax slabs (fallback only —
    'Maharashtra': [(10000, 200), (Decimal('inf'), 200)],  # designation.ed_professionaltax takes priority)
    'Karnataka': [(15000, 200), (Decimal('inf'), 200)],
    'Haryana': [(Decimal('inf'), 0)],    # Haryana currently levies no state PT
}


# =====================================================================
# UTILITY FUNCTIONS
# =====================================================================

def _decimal_round(value):
    """Round to 2 decimal places."""
    return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _get_selected_company(request):
    """Get company from session — matches the pattern used across Aapp modules."""
    cid = request.session.get('selected_company_id')
    return company.objects.filter(company_id=cid).first() if cid else None


def _get_days_in_month(month, year):
    """Return total days in given month."""
    return calendar.monthrange(year, month)[1]


def _get_working_days(month, year):
    """
    Approximate working days excluding Sundays. Used only as a fallback
    when no attendance record exists yet for an employee/month.
    """
    cal = calendar.Calendar()
    working = 0
    for week in cal.monthdayscalendar(year, month):
        working += sum(1 for d in week[:6] if d != 0)  # week[6] is Sunday
    return working


def _get_attendance_summary(employee_obj, month, year):
    """
    Pull the monthly attendance summary record for this employee.
    Aapp's attendance model stores one row per employee per salary_month/
    salary_year with pre-aggregated fields (working_days, leaves, work_pay,
    overtime_hours) — read them directly rather than counting day-level rows.
    Falls back to full working days present if no record has been entered yet.
    """
    from Aapp.app.attandance import attendance
    record = attendance.objects.filter(
        employee_id=employee_obj,
        salary_month=month,
        salary_year=year,
    ).first()

    if record:
        leave_total = (
            record.casual_leaves + record.earned_leaves +
            record.sick_leaves + record.comp_leaves
        )
        # Paid days = days actually worked/on leave-with-pay + work_pay
        # (WP = employee worked on an otherwise unpaid day, e.g. a holiday)
        paid_days = record.working_days + record.work_pay
        return {
            'working_days': record.working_days,
            'paid_days': paid_days,
            'leave_days': leave_total,
            'overtime_hours': record.overtime_hours,
        }

    # No attendance entered yet — default to full working days present
    working = Decimal(_get_working_days(month, year))
    return {
        'working_days': working,
        'paid_days': working,
        'leave_days': Decimal('0'),
        'overtime_hours': Decimal('0'),
    }


def _calculate_pf(basic, pf_applicable, ed_epf_per, er_epf_per):
    """PF from the designation's own rate, subject to the statutory wage ceiling."""
    if not pf_applicable:
        return Decimal('0'), Decimal('0')
    pf_base = min(basic, PF_WAGE_CEILING)
    employee_pf = pf_base * (Decimal(str(ed_epf_per)) / Decimal('100'))
    employer_pf = pf_base * (Decimal(str(er_epf_per)) / Decimal('100'))
    return _decimal_round(employee_pf), _decimal_round(employer_pf)


def _calculate_esi(gross, esi_applicable, ed_esi_per, er_esi_per):
    """ESI from the designation's own rate, only if gross is within the ceiling."""
    if not esi_applicable or gross > ESI_WAGE_CEILING:
        return Decimal('0'), Decimal('0')
    employee_esi = gross * (Decimal(str(ed_esi_per)) / Decimal('100'))
    employer_esi = gross * (Decimal(str(er_esi_per)) / Decimal('100'))
    return _decimal_round(employee_esi), _decimal_round(employer_esi)


def _calculate_labour_welfare(ed_lw_per, er_lw_per, gross):
    """Labour Welfare Fund contribution from the designation's own rate."""
    employee_lw = gross * (Decimal(str(ed_lw_per)) / Decimal('100'))
    employer_lw = gross * (Decimal(str(er_lw_per)) / Decimal('100'))
    return _decimal_round(employee_lw), _decimal_round(employer_lw)


def _calculate_pt(gross, state_name, designation_pt):
    """
    Professional Tax. The designation's own configured rate/amount
    (ed_professionaltax) takes priority; falls back to a state slab lookup
    only if the designation leaves it at 0.
    """
    if designation_pt and designation_pt > 0:
        return Decimal(str(designation_pt))
    slabs = PT_STATES.get(state_name, [])
    for limit, tax in slabs:
        if gross <= limit:
            return Decimal(tax)
    return Decimal('0')


def _calculate_overtime(basic, overtime_hours):
    """Overtime = 2x hourly rate, hourly rate derived from designation's basic pay."""
    if overtime_hours <= 0 or basic <= 0:
        return Decimal('0')
    hourly_basic = basic / Decimal('30') / STANDARD_WORKING_HOURS
    overtime_rate = hourly_basic * Decimal('2')
    return _decimal_round(overtime_rate * overtime_hours)


# =====================================================================
# SALARY CALCULATION ENGINE
# =====================================================================

def calculate_employee_salary(employee_obj, desig, month, year):
    """
    Calculate complete salary for one employee for the given month, using
    their assigned designation's pay scale and statutory rates.

    Returns dict with all earnings, deductions and net pay.
    """
    total_days = Decimal(_get_days_in_month(month, year))

    attendance = _get_attendance_summary(employee_obj, month, year)
    working_days = attendance['working_days']
    paid_days = attendance['paid_days']

    proration = paid_days / working_days if working_days > 0 else Decimal('0')

    # Daily-wage designations pay per day worked directly; salaried
    # designations pay a prorated monthly basic + allowances.
    if desig.is_dailywage:
        daily_rate = Decimal(str(desig.dailywage))
        basic_earned = _decimal_round(daily_rate * paid_days)
        hra_earned = Decimal('0')
        da_earned = Decimal('0')
        conv_earned = Decimal('0')
        medical_earned = Decimal('0')
        lunch_earned = Decimal('0')
        cca_earned = Decimal('0')
        special_earned = Decimal('0')
        travel_earned = Decimal('0')
        washing_earned = Decimal('0')
        cycle_earned = Decimal('0')
        other_earned = Decimal('0')
        basic_for_statutory = daily_rate * working_days  # notional monthly basic for PF ceiling logic
        overtime_amt = _calculate_overtime(daily_rate * Decimal('30'), attendance['overtime_hours'])
    else:
        basic = Decimal(str(desig.basicpay))
        basic_earned = _decimal_round(basic * proration)
        hra_earned = _decimal_round(Decimal(str(desig.hra)) * proration)
        da_earned = _decimal_round(Decimal(str(desig.da)) * proration)
        conv_earned = _decimal_round(Decimal(str(desig.conveyance)) * proration)
        medical_earned = _decimal_round(Decimal(str(desig.medicalallowance)) * proration)
        lunch_earned = _decimal_round(Decimal(str(desig.lunchallowance)) * proration)
        cca_earned = _decimal_round(Decimal(str(desig.cca)) * proration)
        special_earned = _decimal_round(Decimal(str(desig.specialallowance)) * proration)
        travel_earned = _decimal_round(Decimal(str(desig.travelallowance)) * proration)
        washing_earned = _decimal_round(Decimal(str(desig.washingallowance)) * proration)
        cycle_earned = _decimal_round(Decimal(str(desig.cycleallowance)) * proration)
        other_earned = _decimal_round(
            (Decimal(str(desig.other1)) + Decimal(str(desig.other2))) * proration
        )
        basic_for_statutory = basic
        overtime_amt = _calculate_overtime(basic, attendance['overtime_hours'])

    gross = (
        basic_earned + hra_earned + da_earned + conv_earned + medical_earned +
        lunch_earned + cca_earned + special_earned + travel_earned +
        washing_earned + cycle_earned + other_earned + overtime_amt
    )

    # Bonus (from Payment of Bonus Act module, if a record exists for this month)
    bonus_amt = Decimal('0')
    try:
        from Aapp.app.bonus import bonus_record
        bonus_rec = bonus_record.objects.filter(
            employee=employee_obj, salary_month=month, salary_year=year
        ).first()
        if bonus_rec:
            bonus_amt = Decimal(str(bonus_rec.total_bonus))
            gross += bonus_amt
    except Exception:
        pass

    # Company-level statutory registration check — a deduction can only be
    # applied if the company itself is registered under that act. Without
    # this, employee-level enrolment (UAN/ESIC number) alone would let
    # deductions happen even for a company with no EPFO/ESIC/Labour/TAN
    # registration on file, which is not legally valid.
    company_obj = getattr(employee_obj, 'CompanyID', None) or getattr(employee_obj, 'company', None)
    company_epfo_registered = False
    company_esic_registered = False
    company_labour_registered = False
    company_tan_registered = False
    if company_obj is not None:
        try:
            from Sapp.app.company import company_statury
            statury = company_statury.objects.filter(company=company_obj).first()
            if statury:
                company_epfo_registered = bool(statury.epfo)
                company_esic_registered = bool(statury.esic)
                company_labour_registered = bool(statury.labour)
        except Exception:
            pass
        company_tan_registered = bool(getattr(company_obj, 'tan', ''))

    # Statutory deductions — PF/ESI applicability requires BOTH the company
    # being registered under the act AND the employee being enrolled
    # (UAN / ESIC number present on record); rates come from the
    # designation, not a fixed constant.
    pf_applicable = company_epfo_registered and bool(getattr(employee_obj, 'uan_number', ''))
    esi_applicable = company_esic_registered and bool(getattr(employee_obj, 'esic_number', ''))
    lw_applicable = company_labour_registered

    pf_emp, pf_empr = _calculate_pf(
        basic_for_statutory, pf_applicable, desig.ed_epf_per, desig.er_epf_per
    )
    esi_emp, esi_empr = _calculate_esi(
        gross, esi_applicable, desig.ed_esi_per, desig.er_esi_per
    )
    if lw_applicable:
        lw_emp, lw_empr = _calculate_labour_welfare(
            desig.ed_labourwelfare_per, desig.er_labourwelfare_per, gross
        )
    else:
        lw_emp, lw_empr = Decimal('0'), Decimal('0')

    # Professional Tax — designation's own configured amount takes
    # priority; falls back to employee's state slab only if not set.
    state_name = ''
    try:
        state_obj = employee_obj.perm_state or employee_obj.temp_state
        if state_obj:
            state_name = state_obj.name
    except Exception:
        pass
    pt = _calculate_pt(gross, state_name, desig.ed_professionaltax)

    # Income tax withholding requires the company to hold a TAN
    # (Tax Deduction Account Number) — no TAN means no TDS can be filed.
    it = (
        _decimal_round(Decimal(str(desig.ed_income_tax)) * proration)
        if company_tan_registered else Decimal('0')
    )

    total_deductions = pf_emp + esi_emp + lw_emp + pt + it
    net_pay = gross - total_deductions
    employer_cost = gross + pf_empr + esi_empr + lw_empr

    return {
        'total_days': int(total_days),
        'working_days': working_days,
        'paid_days': paid_days,
        'leave_days': attendance['leave_days'],
        'overtime_hours': attendance['overtime_hours'],

        'basic_earned': basic_earned,
        'hra_earned': hra_earned,
        'da_earned': da_earned,
        'conveyance_earned': conv_earned,
        'medical_allowance_earned': medical_earned,
        'lunch_allowance_earned': lunch_earned,
        'cca_earned': cca_earned,
        'special_allowance_earned': special_earned,
        'travel_allowance_earned': travel_earned,
        'washing_allowance_earned': washing_earned,
        'cycle_allowance_earned': cycle_earned,
        'other_allowance_earned': other_earned,
        'overtime_amount': overtime_amt,
        'bonus_amount': bonus_amt,
        'gross_earnings': _decimal_round(gross),

        'pf_deduction': pf_emp,
        'esi_deduction': esi_emp,
        'labour_welfare_deduction': lw_emp,
        'professional_tax': pt,
        'income_tax': it,
        'total_deductions': _decimal_round(total_deductions),

        'net_pay': _decimal_round(net_pay),

        'pf_employer_contribution': pf_empr,
        'esi_employer_contribution': esi_empr,
        'labour_welfare_employer_contribution': lw_empr,
        'employer_total_cost': _decimal_round(employer_cost),
    }


# =====================================================================
# VIEWS - Salary Processing Workflow
# =====================================================================

def salary_dashboard(request):
    """Main salary processing dashboard."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    current_month = timezone.now().month
    current_year = timezone.now().year

    recent_batches = salary_processing.objects.filter(
        company_id=company_obj
    ).order_by('-year', '-month')[:6]

    employees_count = employee.objects.filter(
        CompanyID=company_obj, is_working=True
    ).count()

    # "Structures set" now means: active employees who actually have a
    # designation assigned (pay scale comes from the designation).
    designations_count = designation.objects.filter(
        company=company_obj, is_active=True, is_deleted=False
    ).count()

    context = {
        'company': company_obj,
        'recent_batches': recent_batches,
        'employees_count': employees_count,
        'structures_count': designations_count,
        'month_name': calendar.month_name[current_month],
        'current_year': current_year,
    }
    return render(request, 'Aapp/salary/dashboard.html', context)


def create_salary_batch(request):
    """Initiate salary processing for a month."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    if request.method == 'POST':
        try:
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))

            existing = salary_processing.objects.filter(
                company_id=company_obj, month=month, year=year
            ).first()
            if existing:
                messages.info(
                    request,
                    f'A batch already exists for {dict(MONTH_CHOICES).get(month)} {year}.'
                )
                return redirect('process_salary_batch', batch_id=existing.id)

            batch = salary_processing.objects.create(
                company_id=company_obj,
                month=month,
                year=year,
                status='DRAFT',
                created_by=request.user.username,
            )

            messages.success(
                request,
                f'Salary batch created for {dict(MONTH_CHOICES).get(month)} {year}. '
                'Proceed to process employees.'
            )
            return redirect('process_salary_batch', batch_id=batch.id)

        except Exception as e:
            messages.error(request, f'Error creating batch: {str(e)}')

    today = timezone.now()
    context = {
        'company': company_obj,
        'months': MONTH_CHOICES,
        'current_month': today.month,
        'current_year': today.year,
        'years': YEAR_CHOICES,
    }
    return render(request, 'Aapp/salary/create_batch.html', context)


def process_salary_batch(request, batch_id):
    """
    Process salary for all active employees in the batch, using each
    employee's assigned designation for pay scale and statutory rates.
    """
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    if batch.status == 'APPROVED':
        messages.warning(request, 'Cannot modify approved batch.')
        return redirect('view_salary_batch', batch_id=batch.id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                salary_slip.objects.filter(processing_id=batch).delete()

                employees = employee.objects.filter(
                    CompanyID=company_obj, is_working=True
                ).select_related('perm_state', 'temp_state', 'designationID')

                processed = 0
                skipped = 0
                error_rows = []
                total_gross = Decimal('0')
                total_deductions = Decimal('0')
                total_net = Decimal('0')
                total_employer = Decimal('0')

                for emp in employees:
                    desig = emp.designationID

                    if not desig:
                        skipped += 1
                        error_rows.append(
                            f'{emp.employeecode} - {emp.name}: No designation assigned'
                        )
                        continue

                    try:
                        calc = calculate_employee_salary(
                            emp, desig, batch.month, batch.year
                        )

                        salary_slip.objects.create(
                            processing_id=batch,
                            employee_id=emp,
                            company_id=company_obj,
                            designation_id=desig,
                            **calc,
                            created_by=request.user.username,
                        )

                        total_gross += calc['gross_earnings']
                        total_deductions += calc['total_deductions']
                        total_net += calc['net_pay']
                        total_employer += calc['employer_total_cost']
                        processed += 1

                    except Exception as e:
                        skipped += 1
                        error_rows.append(
                            f'{emp.employeecode} - {emp.name}: {str(e)}'
                        )

                batch.total_employees = processed
                batch.total_gross = _decimal_round(total_gross)
                batch.total_deductions = _decimal_round(total_deductions)
                batch.total_net = _decimal_round(total_net)
                batch.total_employer_cost = _decimal_round(total_employer)
                batch.status = 'PROCESSED'
                batch.processed_at = timezone.now()
                batch.updated_by = request.user.username
                batch.save()

                if error_rows:
                    request.session['salary_errors'] = error_rows

                messages.success(
                    request,
                    f'Salary processed: {processed} employees. '
                    f'Skipped: {skipped}.'
                )
                return redirect('view_salary_batch', batch_id=batch.id)

        except Exception as e:
            messages.error(request, f'Error processing: {str(e)}')
            return redirect('process_salary_batch', batch_id=batch_id)

    slips = salary_slip.objects.filter(processing_id=batch)
    context = {
        'company': company_obj,
        'batch': batch,
        'month_name': dict(MONTH_CHOICES).get(batch.month),
        'slips': slips,
        'employees_count': employee.objects.filter(
            CompanyID=company_obj, is_working=True
        ).count(),
    }
    return render(request, 'Aapp/salary/process_batch.html', context)


def view_salary_batch(request, batch_id):
    """View all salary slips in a batch."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    slips = salary_slip.objects.filter(
        processing_id=batch
    ).select_related('employee_id', 'designation_id').order_by('employee_id__employeecode')

    error_rows = request.session.pop('salary_errors', None)

    context = {
        'company': company_obj,
        'batch': batch,
        'month_name': dict(MONTH_CHOICES).get(batch.month),
        'slips': slips,
        'errors': error_rows,
    }
    return render(request, 'Aapp/salary/view_batch.html', context)


def view_salary_slip(request, slip_id):
    """View individual employee salary slip with full breakdown."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    slip = salary_slip.objects.filter(
        id=slip_id, company_id=company_obj
    ).select_related(
        'employee_id', 'employee_id__perm_state',
        'employee_id__perm_district', 'designation_id', 'processing_id'
    ).first()

    if not slip:
        messages.error(request, 'Salary slip not found.')
        return redirect('salary_dashboard')

    context = {
        'company': company_obj,
        'slip': slip,
        'batch': slip.processing_id,
        'month_name': dict(MONTH_CHOICES).get(slip.processing_id.month),
    }
    return render(request, 'Aapp/salary/view_slip.html', context)


def edit_salary_slip(request, slip_id):
    """Manually adjust a salary slip (bonus, loans, fines, ad-hoc deductions)."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    slip = salary_slip.objects.filter(
        id=slip_id, company_id=company_obj
    ).select_related('employee_id', 'processing_id').first()

    if not slip:
        messages.error(request, 'Salary slip not found.')
        return redirect('salary_dashboard')

    if slip.processing_id.status == 'APPROVED':
        messages.warning(request, 'Cannot edit a slip in an approved batch.')
        return redirect('view_salary_slip', slip_id=slip.id)

    if request.method == 'POST':
        try:
            slip.bonus_amount = Decimal(request.POST.get('bonus_amount', 0) or 0)
            slip.other_allowance_earned = Decimal(request.POST.get('other_allowance_earned', 0) or 0)
            slip.loan_deduction = Decimal(request.POST.get('loan_deduction', 0) or 0)
            slip.advance_deduction = Decimal(request.POST.get('advance_deduction', 0) or 0)
            slip.late_fine = Decimal(request.POST.get('late_fine', 0) or 0)
            slip.other_deduction = Decimal(request.POST.get('other_deduction', 0) or 0)

            slip.gross_earnings = _decimal_round(
                slip.basic_earned + slip.hra_earned + slip.da_earned +
                slip.conveyance_earned + slip.medical_allowance_earned +
                slip.lunch_allowance_earned + slip.cca_earned +
                slip.special_allowance_earned + slip.travel_allowance_earned +
                slip.washing_allowance_earned + slip.cycle_allowance_earned +
                slip.other_allowance_earned + slip.overtime_amount + slip.bonus_amount
            )
            slip.total_deductions = _decimal_round(
                slip.pf_deduction + slip.esi_deduction + slip.labour_welfare_deduction +
                slip.professional_tax + slip.income_tax + slip.loan_deduction +
                slip.advance_deduction + slip.late_fine + slip.other_deduction
            )
            slip.net_pay = _decimal_round(slip.gross_earnings - slip.total_deductions)
            slip.employer_total_cost = _decimal_round(
                slip.gross_earnings + slip.pf_employer_contribution +
                slip.esi_employer_contribution + slip.labour_welfare_employer_contribution
            )
            slip.updated_by = request.user.username
            slip.save()

            messages.success(request, 'Salary slip updated.')
            return redirect('view_salary_slip', slip_id=slip.id)

        except Exception as e:
            messages.error(request, f'Error updating slip: {str(e)}')

    context = {
        'company': company_obj,
        'slip': slip,
        'batch': slip.processing_id,
        'month_name': dict(MONTH_CHOICES).get(slip.processing_id.month),
    }
    return render(request, 'Aapp/salary/edit_slip.html', context)


def approve_salary_batch(request, batch_id):
    """Approve a processed batch, locking it from further edits."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    if batch.status != 'PROCESSED':
        messages.warning(request, 'Only processed batches can be approved.')
        return redirect('view_salary_batch', batch_id=batch.id)

    if request.method == 'POST':
        batch.status = 'APPROVED'
        batch.approved_at = timezone.now()
        batch.updated_by = request.user.username
        batch.save()

        messages.success(request, 'Salary batch approved successfully.')

    return redirect('view_salary_batch', batch_id=batch.id)


def delete_salary_batch(request, batch_id):
    """Delete a salary batch (only if not yet approved)."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    if batch.status == 'APPROVED':
        messages.warning(request, 'Cannot delete an approved batch.')
        return redirect('view_salary_batch', batch_id=batch.id)

    if request.method == 'POST':
        month_name = dict(MONTH_CHOICES).get(batch.month)
        year = batch.year
        batch.delete()
        messages.success(request, f'Salary batch for {month_name} {year} deleted.')
        return redirect('salary_dashboard')

    context = {
        'company': company_obj,
        'batch': batch,
        'month_name': dict(MONTH_CHOICES).get(batch.month),
    }
    return render(request, 'Aapp/salary/delete_batch.html', context)


# =====================================================================
# DESIGNATION PAY SCALES — read-only view over Aapp.app.designation
# =====================================================================
# Pay structure lives entirely on the Designation model (basic, HRA, DA,
# allowances, PF/ESI/PT/IT rates). This module does not duplicate it —
# it reads designations for the active company so an associate can see
# pay scales in the context of salary processing, and links out to the
# real designation management screens (Aapp/app/designation.py) to edit.

def list_salary_structures(request):
    """List designations (pay scales) for the active company."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    designations = designation.objects.filter(
        company=company_obj, is_active=True, is_deleted=False
    ).order_by('designationname')

    # Employee count per designation, for context (annotated onto each object
    # so the template can read d.employee_count directly)
    for desig in designations:
        desig.employee_count = employee.objects.filter(
            CompanyID=company_obj, designationID=desig, is_working=True
        ).count()

    context = {
        'company': company_obj,
        'designations': designations,
    }
    return render(request, 'Aapp/salary/structure_list.html', context)


def create_salary_structure(request):
    """
    Pay scales are set on the Designation, not per-employee. This redirects
    to the real 'add designation' screen rather than duplicating that form.
    """
    messages.info(
        request,
        'Pay scales are configured per designation. Add or edit a designation '
        'to set its pay structure — it will apply to every employee assigned to it.'
    )
    return redirect('create_designation')


def view_salary_structure(request, structure_id):
    """View a single designation's pay scale and statutory rates."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    desig = designation.objects.filter(
        designationid=structure_id, company=company_obj
    ).first()

    if not desig:
        messages.error(request, 'Designation not found.')
        return redirect('list_salary_structures')

    if desig.is_dailywage:
        monthly_gross = Decimal(str(desig.dailywage)) * Decimal('26')
    else:
        monthly_gross = (
            Decimal(str(desig.basicpay)) + Decimal(str(desig.hra)) + Decimal(str(desig.da)) +
            Decimal(str(desig.conveyance)) + Decimal(str(desig.medicalallowance)) +
            Decimal(str(desig.lunchallowance)) + Decimal(str(desig.cca)) +
            Decimal(str(desig.specialallowance)) + Decimal(str(desig.travelallowance)) +
            Decimal(str(desig.washingallowance)) + Decimal(str(desig.cycleallowance)) +
            Decimal(str(desig.other1)) + Decimal(str(desig.other2))
        )
    yearly_gross = monthly_gross * 12

    employee_count = employee.objects.filter(
        CompanyID=company_obj, designationID=desig, is_working=True
    ).count()

    context = {
        'company': company_obj,
        'structure': desig,
        'monthly_gross': _decimal_round(monthly_gross),
        'yearly_gross': _decimal_round(yearly_gross),
        'employee_count': employee_count,
    }
    return render(request, 'Aapp/salary/structure_view.html', context)


# =====================================================================
# EXCEL EXPORT
# =====================================================================

def export_salary_register(request, batch_id):
    """Export complete salary register to Excel."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    slips = salary_slip.objects.filter(
        processing_id=batch
    ).select_related('employee_id').order_by('employee_id__employeecode')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Salary_{calendar.month_name[batch.month]}_{batch.year}'

    hdr_fill = PatternFill('solid', fgColor='1D3557')
    hdr_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:T1')
    ws['A1'] = (
        f'{company_obj.company_name} - Salary Register - '
        f'{calendar.month_name[batch.month]} {batch.year}'
    )
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'Emp Code', 'Name', 'Working Days', 'Paid Days',
        'Basic', 'HRA', 'DA', 'Conveyance', 'Medical', 'Other',
        'Overtime', 'Bonus', 'Gross',
        'PF', 'ESI', 'LWF', 'PT', 'IT', 'Total Ded', 'Net Pay'
    ]

    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 13

    for row_num, slip in enumerate(slips, start=header_row + 1):
        other_allowances_total = float(
            slip.lunch_allowance_earned + slip.cca_earned +
            slip.special_allowance_earned + slip.travel_allowance_earned +
            slip.washing_allowance_earned + slip.cycle_allowance_earned +
            slip.other_allowance_earned
        )
        data = [
            slip.employee_id.employeecode,
            slip.employee_id.name,
            float(slip.working_days),
            float(slip.paid_days),
            float(slip.basic_earned),
            float(slip.hra_earned),
            float(slip.da_earned),
            float(slip.conveyance_earned),
            float(slip.medical_allowance_earned),
            other_allowances_total,
            float(slip.overtime_amount),
            float(slip.bonus_amount),
            float(slip.gross_earnings),
            float(slip.pf_deduction),
            float(slip.esi_deduction),
            float(slip.labour_welfare_deduction),
            float(slip.professional_tax),
            float(slip.income_tax),
            float(slip.total_deductions),
            float(slip.net_pay),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = border
            if col >= 3:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'

    total_row = header_row + len(slips) + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(3, 21):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.cell(
            row=total_row, column=col,
            value=f'=SUM({col_letter}{header_row + 1}:{col_letter}{total_row - 1})'
        ).font = Font(bold=True)
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = (
        f'Salary_{company_obj.company_name.replace(" ", "_")}_'
        f'{batch.month}_{batch.year}.xlsx'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_bank_advice(request, batch_id):
    """Export bank transfer advice file."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batch = salary_processing.objects.filter(
        id=batch_id, company_id=company_obj
    ).first()
    if not batch:
        messages.error(request, 'Salary batch not found.')
        return redirect('salary_dashboard')

    slips = salary_slip.objects.filter(
        processing_id=batch, net_pay__gt=0
    ).select_related('employee_id').order_by('employee_id__employeecode')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bank_Advice'

    headers = ['Emp Code', 'Name', 'Bank Name', 'Account No', 'IFSC', 'Net Pay']
    hdr_fill = PatternFill('solid', fgColor='1D3557')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws.column_dimensions[cell.column_letter].width = 18

    total = Decimal('0')
    for row_num, slip in enumerate(slips, start=2):
        emp = slip.employee_id
        ws.cell(row=row_num, column=1, value=emp.employeecode)
        ws.cell(row=row_num, column=2, value=emp.name)
        ws.cell(row=row_num, column=3, value=emp.bank_name)
        ws.cell(row=row_num, column=4, value=emp.bank_account)
        ws.cell(row=row_num, column=5, value=emp.bank_ifsc)
        ws.cell(row=row_num, column=6, value=float(slip.net_pay)).number_format = '#,##0.00'
        total += slip.net_pay

    total_row = len(slips) + 2
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(total)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="BankAdvice_{batch.month}_{batch.year}.xlsx"'
    )
    wb.save(response)
    return response


def download_salary_template(request):
    """
    Placeholder kept for URL/template compatibility. Pay scales are managed
    on the Designation, not via per-employee Excel import — redirects to
    the designation list where pay scales are actually edited.
    """
    messages.info(
        request,
        'Pay scales are set per designation, not imported per employee. '
        'Edit a designation to change its pay structure for every employee assigned to it.'
    )
    return redirect('list_salary_structures')


def import_salary_structures(request):
    """
    Placeholder kept for URL/template compatibility (structure_list.html's
    bulk-import control posts here). Since pay scales live on Designation,
    there is nothing to import per-employee — redirect back with guidance.
    """
    messages.info(
        request,
        'Pay scales are set per designation. Use the Designations screen to '
        'add or bulk-configure pay structures instead of an Excel import.'
    )
    return redirect('list_salary_structures')


# =====================================================================
# STATUTORY REPORTS
# =====================================================================

def pf_report(request):
    """Generate PF calculation report for the month."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))

    batch = salary_processing.objects.filter(
        company_id=company_obj, month=month, year=year
    ).first()

    slips = []
    totals = {
        'pf_employee': Decimal('0'), 'pf_employer': Decimal('0'),
        'gross': Decimal('0'),
    }

    if batch:
        slips_data = salary_slip.objects.filter(
            processing_id=batch
        ).select_related('employee_id').order_by('employee_id__employeecode')

        for s in slips_data:
            if s.pf_deduction > 0 or s.pf_employer_contribution > 0:
                slips.append(s)
                totals['pf_employee'] += s.pf_deduction
                totals['pf_employer'] += s.pf_employer_contribution
                totals['gross'] += s.gross_earnings

    context = {
        'company': company_obj,
        'slips': slips,
        'totals': {k: _decimal_round(v) for k, v in totals.items()},
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'pf_employee_rate': 'per designation',
        'pf_employer_rate': 'per designation',
    }
    return render(request, 'Aapp/salary/pf_report.html', context)


def esi_report(request):
    """Generate ESI calculation report for the month."""
    company_obj = _get_selected_company(request)
    if not company_obj:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))

    batch = salary_processing.objects.filter(
        company_id=company_obj, month=month, year=year
    ).first()

    slips = []
    totals = {
        'esi_employee': Decimal('0'), 'esi_employer': Decimal('0'),
        'gross': Decimal('0'),
    }

    if batch:
        slips_data = salary_slip.objects.filter(
            processing_id=batch, esi_deduction__gt=0
        ).select_related('employee_id').order_by('employee_id__employeecode')

        for s in slips_data:
            slips.append(s)
            totals['esi_employee'] += s.esi_deduction
            totals['esi_employer'] += s.esi_employer_contribution
            totals['gross'] += s.gross_earnings

    context = {
        'company': company_obj,
        'slips': slips,
        'totals': {k: _decimal_round(v) for k, v in totals.items()},
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'esi_employee_rate': 'per designation',
        'esi_employer_rate': 'per designation',
    }
    return render(request, 'Aapp/salary/esi_report.html', context)