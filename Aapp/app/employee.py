from django import forms
from django.db import models
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from Sapp.app.company import Company
from Sapp.app.state_district import State, District
from Sapp.app.bank import bank_name
from Aapp.app.designation import designation
from Aapp.app.branch_department import branch, department
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime, date

class employee(models.Model):
    # Personal data
    employeeid = models.AutoField(primary_key=True)
    employeecode = models.CharField(max_length=20, unique=True)
    name = models.CharField(max_length=255)
    fathername = models.CharField(max_length=255, blank=True)
    mothername = models.CharField(max_length=255, blank=True)
    gender = models.CharField(max_length=10, choices=[('Male', 'Male'), ('Female', 'Female'), ('Other', 'Other')])
    dateofbirth = models.DateField()
    bloodgroup = models.CharField(max_length=10, blank=True)
    religion = models.CharField(max_length=255, blank=True)
    maritalstatus = models.CharField(max_length=20, choices=[('Single', 'Single'), ('Married', 'Married'), ('Divorced', 'Divorced'), ('Widowed', 'Widowed')], blank=True)
    
    # Temporary Address
    temporaryaddress = models.TextField()
    temp_state = models.ForeignKey(State, on_delete=models.SET_NULL, null=True, related_name='temp_employees')
    temp_district = models.ForeignKey(District, on_delete=models.SET_NULL, null=True, related_name='temp_employees')
    temp_pincode = models.CharField(max_length=10)
    
    # Permanent Address
    same_as_permanent = models.BooleanField(default=False)
    permanentaddress = models.TextField(blank=True)
    perm_state = models.ForeignKey(State, on_delete=models.SET_NULL, null=True, blank=True, related_name='perm_employees')
    perm_district = models.ForeignKey(District, on_delete=models.SET_NULL, null=True, blank=True, related_name='perm_employees')
    perm_pincode = models.CharField(max_length=10, blank=True)
    
    country = models.CharField(max_length=255, default='India')
    mobile = models.CharField(max_length=20, unique=True)
    phone = models.CharField(max_length=20, blank=True)
    email = models.EmailField()
    
    # KYC Fields
    aadhar_number = models.CharField(max_length=12, db_column='Aadhar_Number', unique=True)
    aadhar_name = models.CharField(max_length=200, db_column='Aadhar_Name')
    pan_name = models.CharField(max_length=200, db_column='PAN_Name', blank=True)
    pan_number = models.CharField(max_length=10, db_column='PAN_Number', blank=True, unique=True, null=True)
    bank_name = models.CharField(max_length=255, db_column="Account_Name")
    bank_account = models.CharField(max_length=20, db_column="Account_number", unique=True)
    bank_ifsc = models.CharField(max_length=11, db_column="Bank_IFSC")
    passport_number = models.CharField(max_length=20, db_column="Passport_Number", blank=True)
    passport_expiry = models.DateField(db_column="Passport_Expiry", null=True, blank=True)
    passport_name = models.CharField(max_length=200, db_column="Passport_Name", blank=True)
    driving_license_number = models.CharField(max_length=20, db_column="Driving_License_Number", blank=True)
    driving_license_expiry = models.DateField(db_column="Driving_License_Expiry", null=True, blank=True)
    driving_license_name = models.CharField(max_length=200, db_column="Driving_License_Name", blank=True)
    voterid = models.CharField(max_length=20, db_column="VoterID", blank=True)
    voterid_name = models.CharField(max_length=200, db_column="VoterID_Name", blank=True)
    
    # Employment data
    dateofjoining = models.DateField()
    employment_type = models.CharField(max_length=20, choices=[('Permanent', 'Permanent'), ('Contract', 'Contract'), ('Intern', 'Intern')], default='Permanent')
    contract_start_date = models.DateField(null=True, blank=True)
    contract_end_date = models.DateField(null=True, blank=True)
    internship_start_date = models.DateField(null=True, blank=True)
    internship_end_date = models.DateField(null=True, blank=True)
    uan_number = models.CharField(max_length=50, db_column="UAN", blank=True)
    uan_doj = models.DateField(db_column="EPF_Joining", null=True, blank=True)
    epf_memberID = models.CharField(max_length=100, db_column="EPF_MemberID", blank=True)
    epf_higher = models.BooleanField(default=False)
    esic_number = models.CharField(max_length=10, db_column="ESIC", blank=True)
    esic_doj = models.DateField(db_column="ESI_Joining", null=True, blank=True)
    esic_dol_reason = models.CharField(max_length=255, db_column="ESI_Reason", blank=True)
    labour_id = models.CharField(max_length=50, db_column="LabourID", blank=True)
    dateofretirement = models.DateField(null=True, blank=True)
    dateofleaving = models.DateField(null=True, blank=True)
    leaving_reason = models.CharField(max_length=255, db_column="Leaving_Reason", blank=True)
    is_working = models.BooleanField(default=True)
    
    # Foreign keys
    designationID = models.ForeignKey(designation, on_delete=models.PROTECT, db_column="DesignationID")
    departmentID = models.ForeignKey(department, on_delete=models.PROTECT, db_column="DepartmentID")
    branchID = models.ForeignKey(branch, on_delete=models.PROTECT, db_column="BranchID")
    CompanyID = models.ForeignKey(Company, on_delete=models.PROTECT, db_column='CompanyID')
    
    class Meta:
        db_table = 'employee'
        ordering = ['employeeid']

    def clean(self):
        from django.core.exceptions import ValidationError
        super().clean()
        errors = {}
        today = date.today()
        if self.dateofjoining and self.dateofjoining > today:
            errors['dateofjoining'] = 'Date of joining cannot be in the future.'
        if self.dateofbirth and self.dateofbirth > today:
            errors['dateofbirth'] = 'Date of birth cannot be in the future.'
        if self.dateofbirth and self.dateofjoining and self.dateofjoining < self.dateofbirth:
            errors['dateofjoining'] = 'Date of joining cannot be before date of birth.'
        if self.dateofbirth and self.dateofjoining:
            age_at_joining = (self.dateofjoining - self.dateofbirth).days / 365.25
            if age_at_joining < 14:
                errors['dateofjoining'] = 'Employee must be at least 14 years old as of date of joining.'
        if errors:
            raise ValidationError(errors)

    def __str__(self):
        return f"{self.employeecode} - {self.name}"

# Forms
class EmployeeCompleteForm(forms.ModelForm):
    class Meta:
        model = employee
        fields = '__all__'
        widgets = {
            'dateofbirth': forms.DateInput(attrs={'type': 'date'}),
            'dateofjoining': forms.DateInput(attrs={'type': 'date'}),
            'uan_doj': forms.DateInput(attrs={'type': 'date'}),
            'esic_doj': forms.DateInput(attrs={'type': 'date'}),
            'dateofretirement': forms.DateInput(attrs={'type': 'date'}),
            'dateofleaving': forms.DateInput(attrs={'type': 'date'}),
        }

class EmployeeQuickForm(forms.ModelForm):
    class Meta:
        model = employee
        fields = ['employeecode', 'name', 'gender', 'dateofbirth', 'mobile', 'email', 
                  'aadhar_number', 'aadhar_name', 'bank_name', 'bank_account', 'bank_ifsc',
                  'dateofjoining', 'designationID', 'departmentID', 'branchID', 'CompanyID']
        widgets = {
            'dateofbirth': forms.DateInput(attrs={'type': 'date'}),
            'dateofjoining': forms.DateInput(attrs={'type': 'date'}),
        }

class EmployeeSelfUpdateForm(forms.ModelForm):
    class Meta:
        model = employee
        fields = ['mobile', 'phone', 'email', 'temporaryaddress', 'temp_state', 'temp_district', 
                  'temp_pincode', 'permanentaddress', 'perm_state', 'perm_district', 'perm_pincode',
                  'bank_name', 'bank_account', 'bank_ifsc']

class BulkEmployeeUploadForm(forms.Form):
    csv_file = forms.FileField(label='Upload CSV File', help_text='Upload employee data in CSV format')



# ── helpers ──────────────────────────────────────────────────────────────────
def _company_ctx(request):
    """Return selected company or None."""
    cid = request.session.get('selected_company_id')
    if not cid:
        return None
    from Sapp.app.company import Company
    return Company.objects.filter(company_id=cid).first()


def _form_ctx(company):
    """Return dropdowns scoped to selected company."""
    return {
        'designations': designation.objects.filter(company=company, is_active=True, is_deleted=False),
        'branches':     branch.objects.filter(companyid=company),
        'departments':  department.objects.filter(companyid=company),
        'states':       State.objects.all(),
        'banks':        bank_name.objects.all().order_by('name'),
    }


# ── list ─────────────────────────────────────────────────────────────────────
def list_employee(request):
    from django.contrib import messages
    from django.core.paginator import Paginator
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')
    employees = employee.objects.filter(CompanyID=company).select_related(
        'designationID', 'branchID', 'departmentID').order_by('name')
    paginator = Paginator(employees, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'Aapp/employees/list_employee.html',
                  {'employees': page_obj, 'page_obj': page_obj, 'company': company})


# ── create ────────────────────────────────────────────────────────────────────
def create_employee(request):
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    ctx = _form_ctx(company)
    ctx['company'] = company

    if request.method == 'POST':
        p = request.POST
        required = ['employeecode', 'name', 'gender', 'dateofbirth', 'mobile',
                    'email', 'aadhar_number', 'aadhar_name', 'bank_name',
                    'bank_account', 'bank_ifsc', 'dateofjoining',
                    'designationID', 'departmentID', 'branchID',
                    'temporaryaddress', 'temp_state', 'temp_district', 'temp_pincode']
        if not all(p.get(f) for f in required):
            messages.error(request, 'All required fields must be filled.')
        elif employee.objects.filter(employeecode=p['employeecode'], CompanyID=company).exists():
            messages.error(request, f"Employee code '{p['employeecode']}' already exists.")
        elif employee.objects.filter(mobile=p['mobile']).exists():
            messages.error(request, 'Mobile number already registered.')
        elif employee.objects.filter(aadhar_number=p['aadhar_number']).exists():
            messages.error(request, 'Aadhaar number already registered.')
        elif p.get('pan_number') and employee.objects.filter(pan_number=p['pan_number']).exists():
            messages.error(request, 'PAN number already registered.')
        elif employee.objects.filter(bank_account=p['bank_account']).exists():
            messages.error(request, 'Bank account already registered.')
        else:
            try:
                branch_obj = branch.objects.filter(branchid=p['branchID'], companyid=company).first()
                dept_obj = department.objects.filter(departmentid=p['departmentID'], companyid=company).first()
                designation_obj = designation.objects.filter(
                    designationid=p['designationID'], company=company, is_active=True, is_deleted=False
                ).first()

                if not branch_obj:
                    messages.error(request, 'Branch does not belong to the selected company.')
                    return render(request, 'Aapp/employees/create_employee.html', ctx)
                if not dept_obj:
                    messages.error(request, 'Department does not belong to the selected company.')
                    return render(request, 'Aapp/employees/create_employee.html', ctx)
                if not designation_obj:
                    messages.error(request, 'Designation does not belong to the selected company.')
                    return render(request, 'Aapp/employees/create_employee.html', ctx)
                if dept_obj.branch_id != branch_obj.branchid:
                    messages.error(request, 'Selected department does not belong to the selected branch.')
                    return render(request, 'Aapp/employees/create_employee.html', ctx)

                same = p.get('same_as_permanent') == 'on'
                employee.objects.create(
                    employeecode   = p['employeecode'],
                    name           = p['name'],
                    fathername     = p.get('fathername', ''),
                    mothername     = p.get('mothername', ''),
                    gender         = p['gender'],
                    dateofbirth    = p['dateofbirth'],
                    bloodgroup     = p.get('bloodgroup', ''),
                    religion       = p.get('religion', ''),
                    maritalstatus  = p.get('maritalstatus', ''),
                    temporaryaddress = p['temporaryaddress'],
                    temp_state_id  = p['temp_state'],
                    temp_district_id = p['temp_district'],
                    temp_pincode   = p['temp_pincode'],
                    same_as_permanent = same,
                    permanentaddress = p['temporaryaddress'] if same else p.get('permanentaddress', ''),
                    perm_state_id  = p['temp_state'] if same else (p.get('perm_state') or None),
                    perm_district_id = p['temp_district'] if same else (p.get('perm_district') or None),
                    perm_pincode   = p['temp_pincode'] if same else p.get('perm_pincode', ''),
                    country        = p.get('country', 'India'),
                    mobile         = p['mobile'],
                    phone          = p.get('phone', ''),
                    email          = p['email'],
                    aadhar_number  = p['aadhar_number'],
                    aadhar_name    = p['aadhar_name'],
                    pan_name       = p.get('pan_name', ''),
                    pan_number     = p.get('pan_number') or None,
                    bank_name      = p['bank_name'],
                    bank_account   = p['bank_account'],
                    bank_ifsc      = p['bank_ifsc'],
                    driving_license_name   = p.get('driving_license_name', ''),
                    driving_license_number = p.get('driving_license_number', ''),
                    driving_license_expiry = p.get('driving_license_expiry') or None,
                    passport_name   = p.get('passport_name', ''),
                    passport_number = p.get('passport_number', ''),
                    passport_expiry = p.get('passport_expiry') or None,
                    dateofjoining  = p['dateofjoining'],
                    uan_number     = p.get('uan_number', ''),
                    uan_doj        = p.get('uan_doj') or None,
                    epf_memberID   = p.get('epf_memberID', ''),
                    epf_higher     = p.get('epf_higher') == 'on',
                    esic_number    = p.get('esic_number', ''),
                    esic_doj       = p.get('esic_doj') or None,
                    labour_id      = p.get('labour_id', ''),
                    designationID_id = designation_obj.designationid,
                    departmentID_id  = dept_obj.departmentid,
                    branchID_id      = branch_obj.branchid,
                    CompanyID        = company,
                )
                messages.success(request, f"Employee '{p['name']}' created successfully.")
                return redirect('list_employee')
            except Exception as e:
                messages.error(request, f'Error: {e}')

    return render(request, 'Aapp/employees/create_employee.html', ctx)


# ── alter ─────────────────────────────────────────────────────────────────────
def alter_employee(request, employee_id):
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    emp = get_object_or_404(employee, employeeid=employee_id, CompanyID=company)
    ctx = _form_ctx(company)
    ctx.update({'emp': emp, 'company': company})

    if request.method == 'POST':
        p = request.POST
        try:
            branch_obj = branch.objects.filter(branchid=p.get('branchID'), companyid=company).first()
            dept_obj = department.objects.filter(departmentid=p.get('departmentID'), companyid=company).first()
            designation_obj = designation.objects.filter(
                designationid=p.get('designationID'), company=company, is_active=True, is_deleted=False
            ).first()

            if not branch_obj:
                messages.error(request, 'Branch does not belong to the selected company.')
                return render(request, 'Aapp/employees/alter_employee.html', ctx)
            if not dept_obj:
                messages.error(request, 'Department does not belong to the selected company.')
                return render(request, 'Aapp/employees/alter_employee.html', ctx)
            if not designation_obj:
                messages.error(request, 'Designation does not belong to the selected company.')
                return render(request, 'Aapp/employees/alter_employee.html', ctx)
            if dept_obj.branch_id != branch_obj.branchid:
                messages.error(request, 'Selected department does not belong to the selected branch.')
                return render(request, 'Aapp/employees/alter_employee.html', ctx)

            same = p.get('same_as_permanent') == 'on'
            emp.name           = p.get('name', emp.name)
            emp.fathername     = p.get('fathername', emp.fathername)
            emp.mothername     = p.get('mothername', emp.mothername)
            emp.gender         = p.get('gender', emp.gender)
            emp.dateofbirth    = p.get('dateofbirth', emp.dateofbirth)
            emp.bloodgroup     = p.get('bloodgroup', emp.bloodgroup)
            emp.religion       = p.get('religion', emp.religion)
            emp.maritalstatus  = p.get('maritalstatus', emp.maritalstatus)
            emp.temporaryaddress = p.get('temporaryaddress', emp.temporaryaddress)
            emp.temp_state_id  = p.get('temp_state', emp.temp_state_id)
            emp.temp_district_id = p.get('temp_district', emp.temp_district_id)
            emp.temp_pincode   = p.get('temp_pincode', emp.temp_pincode)
            emp.same_as_permanent = same
            emp.permanentaddress = p['temporaryaddress'] if same else p.get('permanentaddress', emp.permanentaddress)
            emp.perm_state_id  = p['temp_state'] if same else (p.get('perm_state') or emp.perm_state_id)
            emp.perm_district_id = p['temp_district'] if same else (p.get('perm_district') or emp.perm_district_id)
            emp.perm_pincode   = p['temp_pincode'] if same else p.get('perm_pincode', emp.perm_pincode)
            emp.mobile         = p.get('mobile', emp.mobile)
            emp.phone          = p.get('phone', emp.phone)
            emp.email          = p.get('email', emp.email)
            emp.pan_name       = p.get('pan_name', emp.pan_name)
            emp.pan_number     = p.get('pan_number') or emp.pan_number
            emp.bank_name      = p.get('bank_name', emp.bank_name)
            emp.bank_account   = p.get('bank_account', emp.bank_account)
            emp.bank_ifsc      = p.get('bank_ifsc', emp.bank_ifsc)
            emp.uan_number     = p.get('uan_number', emp.uan_number)
            emp.uan_doj        = p.get('uan_doj') or emp.uan_doj
            emp.epf_memberID   = p.get('epf_memberID', emp.epf_memberID)
            emp.epf_higher     = p.get('epf_higher') == 'on'
            emp.esic_number    = p.get('esic_number', emp.esic_number)
            emp.esic_doj       = p.get('esic_doj') or emp.esic_doj
            emp.labour_id      = p.get('labour_id', emp.labour_id)
            emp.designationID_id = designation_obj.designationid
            emp.departmentID_id  = dept_obj.departmentid
            emp.branchID_id      = branch_obj.branchid
            emp.save()
            messages.success(request, f"Employee '{emp.name}' updated successfully.")
            return redirect('list_employee')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'Aapp/employees/alter_employee.html', ctx)


# ── disable / enable ──────────────────────────────────────────────────────────
def disable_employee(request, employee_id):
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    emp = get_object_or_404(employee, employeeid=employee_id, CompanyID=company)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'disable':
            emp.is_working = False
            emp.save()
            messages.success(request, f"Employee '{emp.name}' disabled.")
        elif action == 'enable':
            emp.is_working = True
            emp.dateofleaving = None
            emp.leaving_reason = ''
            emp.save()
            messages.success(request, f"Employee '{emp.name}' re-enabled.")
        return redirect('list_employee')

    return render(request, 'Aapp/employees/disable_employee.html', {'emp': emp})


# ── retire ────────────────────────────────────────────────────────────────────
def retire_employee(request, employee_id):
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    emp = get_object_or_404(employee, employeeid=employee_id, CompanyID=company)

    if request.method == 'POST':
        emp.dateofretirement = request.POST.get('dateofretirement') or None
        emp.dateofleaving    = request.POST.get('dateofleaving') or None
        emp.leaving_reason   = request.POST.get('leaving_reason', '')
        emp.is_working       = False
        emp.save()
        messages.success(request, f"Employee '{emp.name}' retired/separated.")
        return redirect('list_employee')

    return render(request, 'Aapp/employees/retire_employee.html', {'emp': emp})


# ── delete ────────────────────────────────────────────────────────────────────
def delete_employee(request, employee_id):
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    emp = get_object_or_404(employee, employeeid=employee_id, CompanyID=company)

    if request.method == 'POST':
        name = emp.name
        emp.delete()
        messages.success(request, f"Employee '{name}' deleted permanently.")
        return redirect('list_employee')

    return render(request, 'Aapp/employees/delete_employee.html', {'emp': emp})


# ── excel template download ──────────────────────────────────────────────────

@login_required
def download_employee_template(request):
    company = _company_ctx(request)
    if not company:
        from django.contrib import messages
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    headers = [
        'employeecode*', 'name*', 'gender*', 'dateofbirth*', 'mobile*', 'email*',
        'aadhar_number*', 'aadhar_name*', 'bank_name*', 'bank_account*', 'bank_ifsc*',
        'dateofjoining*', 'designation_name*', 'department_name*', 'branch_name*',
        'temporaryaddress*', 'temp_state_name*', 'temp_district_name*', 'temp_pincode*',
        'fathername', 'mothername', 'bloodgroup', 'religion', 'maritalstatus',
        'phone', 'pan_name', 'pan_number', 'employment_type',
    ]
    
    hdr_fill = PatternFill('solid', fgColor='1D3557')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    notes = [
        ('employeecode*',      'Required. Unique employee code.'),
        ('name*',              'Required. Full name of employee.'),
        ('gender*',            'Required. Male/Female/Other.'),
        ('dateofbirth*',       'Required. Format: YYYY-MM-DD.'),
        ('mobile*',            'Required. 10-digit mobile number.'),
        ('email*',             'Required. Valid email address.'),
        ('aadhar_number*',     'Required. 12-digit Aadhaar number.'),
        ('aadhar_name*',       'Required. Name as per Aadhaar.'),
        ('bank_name*',         'Required. Bank name.'),
        ('bank_account*',      'Required. Bank account number.'),
        ('bank_ifsc*',         'Required. Bank IFSC code.'),
        ('dateofjoining*',     'Required. Format: YYYY-MM-DD.'),
        ('designation_name*',  'Required. Must exist in system.'),
        ('department_name*',   'Required. Must exist in system.'),
        ('branch_name*',       'Required. Must exist in system.'),
        ('temporaryaddress*',  'Required. Current address.'),
        ('temp_state_name*',   'Required. State name.'),
        ('temp_district_name*','Required. District name.Use (131001) for district code if name not found.'),
        ('temp_pincode*',      'Required. 6-digit pincode.'),
        ('employment_type',    'Optional. Permanent/Contract/Intern.'),
    ]
    
    ws2.cell(row=1, column=1, value='Column').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Description').font = Font(bold=True)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    
    for r, (col, desc) in enumerate(notes, 2):
        ws2.cell(row=r, column=1, value=col)
        ws2.cell(row=r, column=2, value=desc)

    # Reference data sheet
    ws3 = wb.create_sheet('Reference_Data')
    ws3.cell(row=1, column=1, value='Designations').font = Font(bold=True)
    ws3.cell(row=1, column=2, value='Departments').font = Font(bold=True)
    ws3.cell(row=1, column=3, value='Branches').font = Font(bold=True)
    ws3.cell(row=1, column=4, value='States').font = Font(bold=True)
    ws3.cell(row=1, column=5, value='Districts').font = Font(bold=True)
    ws3.cell(row=1, column=6, value='Banks').font = Font(bold=True)
    
    designations = designation.objects.filter(company=company, is_active=True, is_deleted=False)
    departments = department.objects.filter(companyid=company)
    branches = branch.objects.filter(companyid=company)
    states = State.objects.all()
    Districts = District.objects.all()
    banks = bank_name.objects.all().order_by('name')
    
    for r, d in enumerate(designations, 2):
        ws3.cell(row=r, column=1, value=d.designationname)
    for r, d in enumerate(departments, 2):
        ws3.cell(row=r, column=2, value=d.department_name)
    for r, b in enumerate(branches, 2):
        ws3.cell(row=r, column=3, value=b.branch_name)
    for r, s in enumerate(states, 2):
        ws3.cell(row=r, column=4, value=s.name)
    for r, d in enumerate(Districts, 2):
        ws3.cell(row=r, column=5, value=d.name)
    for r, b in enumerate(banks, 2):
        ws3.cell(row=r, column=6, value=b.name)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="employee_template_{company.company_name}.xlsx"'
    return resp


# ── bulk excel upload ────────────────────────────────────────────────────────

@login_required
def bulk_excel_upload_Employees(request):
    company = _company_ctx(request)
    if not company:
        from django.contrib import messages
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    if request.method == 'POST':
        xl = request.FILES.get('excel_file')
        if not xl:
            from django.contrib import messages
            messages.error(request, 'No file uploaded.')
            return redirect('bulk_excel_upload')

        try:
            wb = openpyxl.load_workbook(xl, data_only=True)
            ws = wb.active
        except Exception:
            from django.contrib import messages
            messages.error(request, 'Invalid Excel file.')
            return redirect('bulk_excel_upload')

        # Build lookup maps
        designation_map = {d.designationname: d for d in designation.objects.filter(company=company, is_active=True, is_deleted=False)}
        department_map = {d.department_name: d for d in department.objects.filter(companyid=company)}
        branch_map = {b.branch_name: b for b in branch.objects.filter(companyid=company)}
        state_map = {s.name: s for s in State.objects.all()}
        
        created = skipped = errors = 0
        error_rows = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            # Required fields
            employeecode = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if row[1] else ''
            gender = str(row[2]).strip() if row[2] else ''
            dateofbirth = row[3]
            mobile = str(row[4]).strip() if row[4] else ''
            email = str(row[5]).strip() if row[5] else ''
            aadhar_number = str(row[6]).strip() if row[6] else ''
            aadhar_name = str(row[7]).strip() if row[7] else ''
            bank_name_val = str(row[8]).strip() if row[8] else ''
            bank_account = str(row[9]).strip() if row[9] else ''
            bank_ifsc = str(row[10]).strip() if row[10] else ''
            dateofjoining = row[11]
            designation_name = str(row[12]).strip() if row[12] else ''
            department_name = str(row[13]).strip() if row[13] else ''
            branch_name = str(row[14]).strip() if row[14] else ''
            temporaryaddress = str(row[15]).strip() if row[15] else ''
            temp_state_name = str(row[16]).strip() if row[16] else ''
            temp_district_name = str(row[17]).strip() if row[17] else ''
            temp_pincode = str(row[18]).strip() if row[18] else ''
            
            # Optional fields
            fathername = str(row[19]).strip() if row[19] else ''
            mothername = str(row[20]).strip() if row[20] else ''
            bloodgroup = str(row[21]).strip() if row[21] else ''
            religion = str(row[22]).strip() if row[22] else ''
            maritalstatus = str(row[23]).strip() if row[23] else ''
            phone = str(row[24]).strip() if row[24] else ''
            pan_name = str(row[25]).strip() if row[25] else ''
            pan_number = str(row[26]).strip() if row[26] else ''
            employment_type = str(row[27]).strip() if row[27] else 'Permanent'

            # Validate required fields
            required_fields = [
                (employeecode, 'employeecode'), (name, 'name'), (gender, 'gender'),
                (mobile, 'mobile'), (email, 'email'), (aadhar_number, 'aadhar_number'),
                (aadhar_name, 'aadhar_name'), (bank_name_val, 'bank_name'),
                (bank_account, 'bank_account'), (bank_ifsc, 'bank_ifsc'),
                (designation_name, 'designation_name'), (department_name, 'department_name'),
                (branch_name, 'branch_name'), (temporaryaddress, 'temporaryaddress'),
                (temp_state_name, 'temp_state_name'), (temp_district_name, 'temp_district_name'),
                (temp_pincode, 'temp_pincode')
            ]
            
            missing = [field for val, field in required_fields if not val]
            if missing or not dateofbirth or not dateofjoining:
                error_rows.append(f"Row {row_num}: Missing required fields: {', '.join(missing + (['dateofbirth'] if not dateofbirth else []) + (['dateofjoining'] if not dateofjoining else []))}")
                errors += 1
                continue

            # Validate gender
            if gender not in ['Male', 'Female', 'Other']:
                error_rows.append(f"Row {row_num}: Gender must be Male/Female/Other.")
                errors += 1
                continue

            # Validate employment type
            if employment_type not in ['Permanent', 'Contract', 'Intern']:
                employment_type = 'Permanent'

            # Parse dates
            try:
                if isinstance(dateofbirth, str):
                    dateofbirth = datetime.strptime(dateofbirth, '%Y-%m-%d').date()
                if isinstance(dateofjoining, str):
                    dateofjoining = datetime.strptime(dateofjoining, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                error_rows.append(f"Row {row_num}: Invalid date format. Use YYYY-MM-DD.")
                errors += 1
                continue

            today = date.today()
            if dateofjoining > today:
                error_rows.append(f"Row {row_num}: Date of joining ({dateofjoining}) cannot be in the future.")
                errors += 1
                continue
            if dateofbirth > today:
                error_rows.append(f"Row {row_num}: Date of birth ({dateofbirth}) cannot be in the future.")
                errors += 1
                continue
            if dateofjoining < dateofbirth:
                error_rows.append(f"Row {row_num}: Date of joining cannot be before date of birth.")
                errors += 1
                continue
            age_at_joining = (dateofjoining - dateofbirth).days / 365.25
            if age_at_joining < 14:
                error_rows.append(f"Row {row_num}: Employee must be at least 14 years old as of date of joining.")
                errors += 1
                continue

            # Check duplicates
            if employee.objects.filter(employeecode=employeecode, CompanyID=company).exists():
                error_rows.append(f"Row {row_num}: Employee code '{employeecode}' already exists.")
                errors += 1
                continue
            if employee.objects.filter(mobile=mobile).exists():
                error_rows.append(f"Row {row_num}: Mobile '{mobile}' already registered.")
                errors += 1
                continue
            if employee.objects.filter(aadhar_number=aadhar_number).exists():
                error_rows.append(f"Row {row_num}: Aadhaar '{aadhar_number}' already registered.")
                errors += 1
                continue
            if employee.objects.filter(bank_account=bank_account).exists():
                error_rows.append(f"Row {row_num}: Bank account '{bank_account}' already registered.")
                errors += 1
                continue
            if pan_number and employee.objects.filter(pan_number=pan_number).exists():
                error_rows.append(f"Row {row_num}: PAN '{pan_number}' already registered.")
                errors += 1
                continue

            # Lookup foreign keys
            desig = designation_map.get(designation_name)
            dept = department_map.get(department_name)
            br = branch_map.get(branch_name)
            state = state_map.get(temp_state_name)
            
            if not desig:
                error_rows.append(f"Row {row_num}: Designation '{designation_name}' not found.")
                errors += 1
                continue
            if not dept:
                error_rows.append(f"Row {row_num}: Department '{department_name}' not found.")
                errors += 1
                continue
            if not br:
                error_rows.append(f"Row {row_num}: Branch '{branch_name}' not found.")
                errors += 1
                continue
            if not state:
                error_rows.append(f"Row {row_num}: State '{temp_state_name}' not found.")
                errors += 1
                continue

            # Find district
            district = District.objects.filter(state=state, name=temp_district_name).first()
            if not district:
                error_rows.append(f"Row {row_num}: District '{temp_district_name}' not found in state '{temp_state_name}'.")
                errors += 1
                continue

            try:
                employee.objects.create(
                    employeecode=employeecode,
                    name=name,
                    fathername=fathername,
                    mothername=mothername,
                    gender=gender,
                    dateofbirth=dateofbirth,
                    bloodgroup=bloodgroup,
                    religion=religion,
                    maritalstatus=maritalstatus,
                    temporaryaddress=temporaryaddress,
                    temp_state=state,
                    temp_district=district,
                    temp_pincode=temp_pincode,
                    permanentaddress=temporaryaddress,
                    perm_state=state,
                    perm_district=district,
                    perm_pincode=temp_pincode,
                    country='India',
                    mobile=mobile,
                    phone=phone,
                    email=email,
                    aadhar_number=aadhar_number,
                    aadhar_name=aadhar_name,
                    pan_name=pan_name,
                    pan_number=pan_number or None,
                    bank_name=bank_name_val,
                    bank_account=bank_account,
                    bank_ifsc=bank_ifsc,
                    dateofjoining=dateofjoining,
                    employment_type=employment_type,
                    designationID=desig,
                    departmentID=dept,
                    branchID=br,
                    CompanyID=company,
                )
                created += 1
            except Exception as e:
                error_rows.append(f"Row {row_num}: {e}")
                errors += 1

        from django.contrib import messages
        if created:
            messages.success(request, f"{created} employee(s) imported, {errors} error(s).")
        else:
            messages.warning(request, f"No employees imported. {errors} error(s).")

        if error_rows:
            for err in error_rows[:10]:  # show max 10 errors
                messages.error(request, err)

        return redirect('list_employee')

    return render(request, 'Aapp/employees/bulk_excel_upload.html', {'company': company})


# ── bulk statutory fields update (UAN/EPF, ESIC, PAN, Labour ID) ─────────────
def bulk_update_statutory_fields(request):
    """
    Show every active employee missing UAN/EPF Member ID, ESIC Number,
    PAN Number, or Labour ID in an editable table, and save all entered
    values in a single submit. Only employees with at least one of these
    four fields blank are listed; fields left blank in the form keep
    their existing value (no accidental clearing).
    """
    from django.contrib import messages
    company = _company_ctx(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    if request.method == 'POST':
        emp_ids = request.POST.getlist('employee_id')
        updated = 0
        errors = 0
        error_rows = []

        # Guard against two rows in the same submit claiming the same PAN
        submitted_pans = {}

        with_transaction = True
        from django.db import transaction as _txn
        try:
            with _txn.atomic():
                for emp_id in emp_ids:
                    emp = employee.objects.filter(
                        employeeid=emp_id, CompanyID=company
                    ).first()
                    if not emp:
                        continue

                    uan_number = request.POST.get(f'uan_number_{emp_id}', '').strip()
                    epf_memberID = request.POST.get(f'epf_memberID_{emp_id}', '').strip()
                    esic_number = request.POST.get(f'esic_number_{emp_id}', '').strip()
                    pan_number = request.POST.get(f'pan_number_{emp_id}', '').strip().upper()
                    labour_id = request.POST.get(f'labour_id_{emp_id}', '').strip()

                    # Nothing entered for this row — skip silently
                    if not any([uan_number, epf_memberID, esic_number, pan_number, labour_id]):
                        continue

                    try:
                        if pan_number:
                            if pan_number in submitted_pans:
                                error_rows.append(
                                    f"{emp.employeecode} - {emp.name}: PAN '{pan_number}' "
                                    f"was also entered for {submitted_pans[pan_number]} in this batch."
                                )
                                errors += 1
                                continue
                            existing_pan = employee.objects.filter(
                                pan_number=pan_number
                            ).exclude(employeeid=emp.employeeid).first()
                            if existing_pan:
                                error_rows.append(
                                    f"{emp.employeecode} - {emp.name}: PAN '{pan_number}' "
                                    f"already registered to {existing_pan.employeecode} - {existing_pan.name}."
                                )
                                errors += 1
                                continue
                            submitted_pans[pan_number] = f"{emp.employeecode} - {emp.name}"
                            emp.pan_number = pan_number
                            if not emp.pan_name:
                                emp.pan_name = emp.name

                        if uan_number:
                            emp.uan_number = uan_number
                        if epf_memberID:
                            emp.epf_memberID = epf_memberID
                        if esic_number:
                            emp.esic_number = esic_number
                        if labour_id:
                            emp.labour_id = labour_id

                        emp.save()
                        updated += 1

                    except Exception as e:
                        error_rows.append(f"{emp.employeecode} - {emp.name}: {e}")
                        errors += 1

        except Exception as e:
            messages.error(request, f'Error saving updates: {e}')
            return redirect('bulk_update_statutory_fields')

        if updated:
            messages.success(request, f"{updated} employee record(s) updated, {errors} error(s).")
        else:
            messages.warning(request, f"No records updated. {errors} error(s).")

        if error_rows:
            for err in error_rows[:10]:
                messages.error(request, err)

        return redirect('bulk_update_statutory_fields')

    # GET — list employees missing at least one of the four fields
    incomplete = employee.objects.filter(
        CompanyID=company, is_working=True
    ).filter(
        models.Q(uan_number='') | models.Q(epf_memberID='') |
        models.Q(esic_number='') | models.Q(pan_number__isnull=True) |
        models.Q(pan_number='') | models.Q(labour_id='')
    ).select_related('designationID', 'branchID').order_by('employeecode')

    context = {
        'company': company,
        'employees': incomplete,
    }
    return render(request, 'Aapp/employees/bulk_update_statutory.html', context)