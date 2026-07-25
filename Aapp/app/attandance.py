from django import forms
from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.urls import reverse
from Sapp.app.company import Company
from Aapp.app.branch_department import branch
from Aapp.app.employee import employee
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

MONTH_CHOICES = [
    (1,'January'),(2,'February'),(3,'March'),(4,'April'),
    (5,'May'),(6,'June'),(7,'July'),(8,'August'),
    (9,'September'),(10,'October'),(11,'November'),(12,'December'),
]
YEAR_CHOICES = [(y, y) for y in range(2026, 2032)]


# ── Model ─────────────────────────────────────────────────────────────────────

class attendance(models.Model):
    attendanceid  = models.AutoField(primary_key=True)
    is_bulk       = models.BooleanField(default=False)
    employee_id   = models.ForeignKey(employee, on_delete=models.CASCADE, db_column='employee_id', related_name='attendances')
    emp_code      = models.CharField(max_length=20)
    companyid     = models.ForeignKey(Company, on_delete=models.CASCADE, db_column='companyid')
    divisionid    = models.CharField(max_length=100, blank=True, default='')
    branchid      = models.ForeignKey(branch, on_delete=models.SET_NULL, null=True, blank=True, db_column='branchid')
    salary_month  = models.PositiveSmallIntegerField(choices=MONTH_CHOICES)
    salary_year   = models.PositiveSmallIntegerField(choices=YEAR_CHOICES)
    working_days  = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    holidays      = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    casual_leaves = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    earned_leaves = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    sick_leaves   = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    comp_leaves   = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    work_pay      = models.DecimalField(max_digits=5, decimal_places=2, default=0)
    overtime_hours = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    ordinary_rate  = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    ot_rate        = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    ot_wages_paid  = models.DecimalField(max_digits=10, decimal_places=2, default=0)
    created_by    = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, related_name='attendance_created')
    created_date  = models.DateTimeField(auto_now_add=True)
    updated_by    = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, blank=True, related_name='attendance_updated')
    updated_date  = models.DateTimeField(auto_now=True)

    class Meta:
        db_table = 'attendance'
        unique_together = ('employee_id', 'salary_month', 'salary_year')
        ordering = ['-salary_year', '-salary_month']

    def __str__(self):
        return f"{self.emp_code} — {self.get_salary_month_display()} {self.salary_year}"

    def save(self, *args, **kwargs):
        if self.ordinary_rate and not self.ot_rate:
            self.ot_rate = self.ordinary_rate * 2
        if self.overtime_hours and self.ot_rate and not self.ot_wages_paid:
            self.ot_wages_paid = self.overtime_hours * self.ot_rate
        super().save(*args, **kwargs)

class MinimumWagesOvertimeRegister(models.Model):
    ot_register_id = models.AutoField(primary_key=True)
    attendance = models.ForeignKey(attendance, on_delete=models.CASCADE, related_name='ot_details')
    ot_date = models.DateField()
    overtime_hours = models.DecimalField(max_digits=6, decimal_places=2)
    ordinary_rate = models.DecimalField(max_digits=10, decimal_places=2)
    ot_rate = models.DecimalField(max_digits=10, decimal_places=2)
    ot_wages_paid = models.DecimalField(max_digits=10, decimal_places=2)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    class Meta:
        db_table = 'minimum_wages_overtime_register'
        ordering = ['-ot_date']

    def __str__(self):
        return f"{self.attendance.emp_code} - {self.ot_date}"

    def save(self, *args, **kwargs):
        if self.ordinary_rate and not self.ot_rate:
            self.ot_rate = self.ordinary_rate * 2
        if self.overtime_hours and self.ot_rate:
            self.ot_wages_paid = self.overtime_hours * self.ot_rate
        super().save(*args, **kwargs)


# ── Form ──────────────────────────────────────────────────────────────────────

class AttendanceForm(forms.ModelForm):
    class Meta:
        model = attendance
        fields = [
            'employee_id', 'emp_code', 'companyid', 'divisionid', 'branchid',
            'salary_month', 'salary_year', 'working_days', 'holidays',
            'casual_leaves', 'earned_leaves', 'sick_leaves', 'comp_leaves', 'work_pay',
            'overtime_hours', 'ordinary_rate', 'ot_rate', 'ot_wages_paid',
        ]


class MinimumWagesOvertimeRegisterForm(forms.ModelForm):
    class Meta:
        model = MinimumWagesOvertimeRegister
        fields = ['attendance', 'ot_date', 'overtime_hours', 'ordinary_rate', 'ot_rate', 'ot_wages_paid']
        widgets = {
            'ot_date': forms.DateInput(attrs={'type': 'date'}),
        }


# ── Helper ────────────────────────────────────────────────────────────────────

def _company(request):
    cid = request.session.get('selected_company_id')
    return Company.objects.filter(company_id=cid).first() if cid else None


# ── list ──────────────────────────────────────────────────────────────────────

@login_required
def list_attendance(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')
    records = attendance.objects.filter(companyid=company).select_related('employee_id', 'branchid')
    return render(request, 'Aapp/attendance/list_attendance.html',
                  {'records': records, 'company': company})


# ── add ───────────────────────────────────────────────────────────────────────

@login_required
def add_attendance(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    employees = employee.objects.filter(CompanyID=company, is_working=True).order_by('name')
    branches  = branch.objects.filter(companyid=company)

    if request.method == 'POST':
        p     = request.POST
        emp   = get_object_or_404(employee, employeeid=p.get('employee_id'), CompanyID=company)
        month = int(p.get('salary_month', 0))
        year  = int(p.get('salary_year', 0))

        if attendance.objects.filter(employee_id=emp, salary_month=month, salary_year=year).exists():
            messages.error(request, f"Attendance for {emp.name} — {month}/{year} already exists.")
        else:
            try:
                attendance.objects.create(
                    is_bulk       = False,
                    employee_id   = emp,
                    emp_code      = emp.employeecode,
                    companyid     = company,
                    divisionid    = p.get('divisionid', ''),
                    branchid_id   = p.get('branchid') or None,
                    salary_month  = month,
                    salary_year   = year,
                    working_days  = p.get('working_days', 0),
                    holidays      = p.get('holidays', 0),
                    casual_leaves = p.get('casual_leaves', 0),
                    earned_leaves = p.get('earned_leaves', 0),
                    sick_leaves   = p.get('sick_leaves', 0),
                    comp_leaves   = p.get('comp_leaves', 0),
                    work_pay      = p.get('work_pay', 0),
                    overtime_hours = p.get('overtime_hours', 0),
                    ordinary_rate  = p.get('ordinary_rate', 0),
                    ot_rate        = p.get('ot_rate', 0),
                    ot_wages_paid  = p.get('ot_wages_paid', 0),
                    created_by    = request.user,
                )
                messages.success(request, f"Attendance for {emp.name} saved.")
                return redirect('list_attendance')
            except Exception as e:
                messages.error(request, f"Error: {e}")

    return render(request, 'Aapp/attendance/add_attendance.html', {
        'employees': employees, 'branches': branches,
        'months': MONTH_CHOICES, 'company': company,
    })


# ── update ────────────────────────────────────────────────────────────────────

@login_required
def update_attendance(request, attendance_id):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    rec      = get_object_or_404(attendance, attendanceid=attendance_id, companyid=company)
    branches = branch.objects.filter(companyid=company)

    if request.method == 'POST':
        p = request.POST
        try:
            rec.divisionid    = p.get('divisionid', rec.divisionid)
            rec.branchid_id   = p.get('branchid') or rec.branchid_id
            rec.salary_month  = int(p.get('salary_month', rec.salary_month))
            rec.salary_year   = int(p.get('salary_year', rec.salary_year))
            rec.working_days  = p.get('working_days', rec.working_days)
            rec.holidays      = p.get('holidays', rec.holidays)
            rec.casual_leaves = p.get('casual_leaves', rec.casual_leaves)
            rec.earned_leaves = p.get('earned_leaves', rec.earned_leaves)
            rec.sick_leaves   = p.get('sick_leaves', rec.sick_leaves)
            rec.comp_leaves   = p.get('comp_leaves', rec.comp_leaves)
            rec.work_pay      = p.get('work_pay', rec.work_pay)
            rec.overtime_hours = p.get('overtime_hours', rec.overtime_hours)
            rec.ordinary_rate  = p.get('ordinary_rate', rec.ordinary_rate)
            rec.ot_rate        = p.get('ot_rate', rec.ot_rate)
            rec.ot_wages_paid  = p.get('ot_wages_paid', rec.ot_wages_paid)
            rec.updated_by    = request.user
            rec.save()
            messages.success(request, "Attendance updated.")
            return redirect('list_attendance')
        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, 'Aapp/attendance/update_attendance.html', {
        'rec': rec, 'branches': branches, 'months': MONTH_CHOICES,
    })


# ── bulk ──────────────────────────────────────────────────────────────────────

@login_required
def bulk_attendance(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    employees = employee.objects.filter(CompanyID=company, is_working=True).order_by('name')
    branches  = branch.objects.filter(companyid=company)

    if request.method == 'POST':
        p          = request.POST
        month      = int(p.get('salary_month', 0))
        year       = int(p.get('salary_year', 0))
        branch_id  = p.get('branchid') or None
        divisionid = p.get('divisionid', '')
        emp_ids    = p.getlist('employee_ids')
        created = skipped = 0

        for eid in emp_ids:
            emp = employee.objects.filter(employeeid=eid, CompanyID=company).first()
            if not emp:
                continue
            if attendance.objects.filter(employee_id=emp, salary_month=month, salary_year=year).exists():
                skipped += 1
                continue
            attendance.objects.create(
                is_bulk       = True,
                employee_id   = emp,
                emp_code      = emp.employeecode,
                companyid     = company,
                divisionid    = divisionid,
                branchid_id   = branch_id,
                salary_month  = month,
                salary_year   = year,
                working_days  = p.get(f'working_days_{eid}', 0),
                holidays      = p.get(f'holidays_{eid}', 0),
                casual_leaves = p.get(f'casual_leaves_{eid}', 0),
                earned_leaves = p.get(f'earned_leaves_{eid}', 0),
                sick_leaves   = p.get(f'sick_leaves_{eid}', 0),
                comp_leaves   = p.get(f'comp_leaves_{eid}', 0),
                work_pay      = p.get(f'work_pay_{eid}', 0),
                overtime_hours = p.get(f'overtime_hours_{eid}', 0),
                ordinary_rate  = p.get(f'ordinary_rate_{eid}', 0),
                ot_rate        = p.get(f'ot_rate_{eid}', 0),
                ot_wages_paid  = p.get(f'ot_wages_paid_{eid}', 0),
                created_by    = request.user,
            )
            created += 1

        messages.success(request, f"{created} record(s) saved, {skipped} skipped (already exist).")
        return redirect('list_attendance')

    return render(request, 'Aapp/attendance/bulk_attendance.html', {
        'employees': employees, 'branches': branches,
        'months': MONTH_CHOICES, 'company': company,
    })


# ── delete ────────────────────────────────────────────────────────────────────

@login_required
def delete_attendance(request, attendance_id):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    rec = get_object_or_404(attendance, attendanceid=attendance_id, companyid=company)

    if request.method == 'POST':
        rec.delete()
        messages.success(request, "Attendance record deleted.")
        return redirect('list_attendance')

    return render(request, 'Aapp/attendance/delete_attendance.html', {'rec': rec})


# ── excel template download ───────────────────────────────────────────────────

@login_required
def download_attendance_template(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    headers = [
        'emp_code*', 'salary_month*', 'salary_year*',
        'working_days', 'holidays', 'casual_leaves',
        'earned_leaves', 'sick_leaves', 'comp_leaves', 'work_pay',
        'overtime_hours', 'ordinary_rate', 'ot_rate', 'ot_wages_paid',
    ]
    hdr_fill = PatternFill('solid', fgColor='1D3557')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 16

    # Pre-fill emp_code column with active employees
    employees = employee.objects.filter(CompanyID=company, is_working=True).order_by('name').values_list('employeecode', flat=True)
    for row, code in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=code)

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    notes = [
        ('emp_code*',      'Required. Must match an active employee code for this company.'),
        ('salary_month*',  'Required. Integer 1–12.'),
        ('salary_year*',   'Required. 4-digit year e.g. 2025.'),
        ('working_days',   'Decimal. Defaults to 0 if blank.'),
        ('holidays',       'Decimal. Defaults to 0 if blank.'),
        ('casual_leaves',  'Decimal. Defaults to 0 if blank.'),
        ('earned_leaves',  'Decimal. Defaults to 0 if blank.'),
        ('sick_leaves',    'Decimal. Defaults to 0 if blank.'),
        ('comp_leaves',    'Decimal. Defaults to 0 if blank.'),
        ('work_pay',       'Decimal. Defaults to 0 if blank.'),
        ('overtime_hours', 'Decimal. OT hours worked. Defaults to 0 if blank.'),
        ('ordinary_rate',  'Decimal. Regular hourly rate. Defaults to 0 if blank.'),
        ('ot_rate',        'Decimal. OT rate (2× ordinary). Auto-calculated if blank.'),
        ('ot_wages_paid',  'Decimal. OT wages. Auto-calculated if blank.'),
    ]
    ws2.cell(row=1, column=1, value='Column').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Description').font = Font(bold=True)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 60
    for r, (col, desc) in enumerate(notes, 2):
        ws2.cell(row=r, column=1, value=col)
        ws2.cell(row=r, column=2, value=desc)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="attendance_template_{company.company_name}.xlsx"'
    return resp


# ── bulk excel upload ─────────────────────────────────────────────────────────

@login_required
def bulk_excel_upload_Attandance(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    if request.method == 'POST':
        xl = request.FILES.get('excel_file')
        if not xl:
            messages.error(request, 'No file uploaded.')
            return redirect('bulk_excel_upload')

        try:
            wb = openpyxl.load_workbook(xl, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'Invalid Excel file.')
            return redirect('bulk_excel_upload')

        # Build emp_code → employee map for this company
        emp_map = {e.employeecode: e for e in employee.objects.filter(CompanyID=company, is_working=True)}

        created = skipped = errors = 0
        error_rows = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            emp_code     = str(row[0]).strip() if row[0] else ''
            salary_month = row[1]
            salary_year  = row[2]
            working_days  = row[3] or 0
            holidays      = row[4] or 0
            casual_leaves = row[5] or 0
            earned_leaves = row[6] or 0
            sick_leaves   = row[7] or 0
            comp_leaves   = row[8] or 0
            work_pay      = row[9] or 0
            overtime_hours = row[10] or 0
            ordinary_rate  = row[11] or 0
            ot_rate        = row[12] or 0
            ot_wages_paid  = row[13] or 0

            # Validate required fields
            if not emp_code or not salary_month or not salary_year:
                error_rows.append(f"Row {row_num}: emp_code, salary_month, salary_year are required.")
                errors += 1
                continue

            try:
                salary_month = int(salary_month)
                salary_year  = int(salary_year)
            except (ValueError, TypeError):
                error_rows.append(f"Row {row_num}: salary_month and salary_year must be integers.")
                errors += 1
                continue

            if salary_month < 1 or salary_month > 12:
                error_rows.append(f"Row {row_num}: salary_month must be 1–12.")
                errors += 1
                continue

            emp = emp_map.get(emp_code)
            if not emp:
                error_rows.append(f"Row {row_num}: Employee code '{emp_code}' not found or inactive.")
                errors += 1
                continue

            if attendance.objects.filter(employee_id=emp, salary_month=salary_month, salary_year=salary_year).exists():
                skipped += 1
                continue

            try:
                attendance.objects.create(
                    is_bulk       = True,
                    employee_id   = emp,
                    emp_code      = emp.employeecode,
                    companyid     = company,
                    salary_month  = salary_month,
                    salary_year   = salary_year,
                    working_days  = working_days,
                    holidays      = holidays,
                    casual_leaves = casual_leaves,
                    earned_leaves = earned_leaves,
                    sick_leaves   = sick_leaves,
                    comp_leaves   = comp_leaves,
                    work_pay      = work_pay,
                    overtime_hours = overtime_hours,
                    ordinary_rate  = ordinary_rate,
                    ot_rate        = ot_rate,
                    ot_wages_paid  = ot_wages_paid,
                    created_by    = request.user,
                )
                created += 1
            except Exception as e:
                error_rows.append(f"Row {row_num}: {e}")
                errors += 1

        if created:
            messages.success(request, f"{created} record(s) imported, {skipped} skipped (duplicate), {errors} error(s).")
        else:
            messages.warning(request, f"No records imported. {skipped} duplicate(s), {errors} error(s).")

        if error_rows:
            for err in error_rows[:10]:   # show max 10 errors
                messages.error(request, err)

        return redirect('list_attendance')

    return render(request, 'Aapp/attendance/bulk_excel_upload.html', {'company': company})


# ── Overtime Register Views ──────────────────────────────────────────────────

@login_required
def list_overtime_register(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')
    records = MinimumWagesOvertimeRegister.objects.filter(
        attendance__companyid=company
    ).select_related('attendance__employee_id')
    rows = [{
        'cells': [r.attendance.emp_code, r.ot_date, r.overtime_hours, r.ordinary_rate, r.ot_rate, r.ot_wages_paid],
        'actions': [
            {'url': reverse('alter_overtime_register', args=[r.ot_register_id]), 'label': 'Edit', 'css': 'edit'},
            {'url': reverse('delete_overtime_register', args=[r.ot_register_id]), 'label': 'Delete', 'css': 'delete'},
        ],
    } for r in records]
    return render(request, 'Aapp/generic/list.html', {
        'page_title': 'Minimum Wages Act — Overtime Register (Form IV)',
        'columns': ['Employee Code', 'OT Date', 'OT Hours', 'Ordinary Rate', 'OT Rate (2x)', 'OT Wages Paid'],
        'rows': rows, 'company': company,
        'add_url': reverse('create_overtime_register'), 'add_label': 'Add Overtime Record',
        'empty_message': 'No overtime records yet.',
    })


@login_required
def create_overtime_register(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    if request.method == 'POST':
        form = MinimumWagesOvertimeRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Overtime record created successfully.')
            return redirect('list_overtime_register')
    else:
        form = MinimumWagesOvertimeRegisterForm()
        form.fields['attendance'].queryset = attendance.objects.filter(companyid=company)

    return render(request, 'Aapp/generic/form.html', {
        'form': form, 'company': company,
        'page_title': 'Add Overtime Record (Minimum Wages Act — Form IV)',
        'cancel_url': reverse('list_overtime_register'),
    })


@login_required
def alter_overtime_register(request, ot_register_id):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    record = get_object_or_404(MinimumWagesOvertimeRegister, ot_register_id=ot_register_id,
                                attendance__companyid=company)

    if request.method == 'POST':
        form = MinimumWagesOvertimeRegisterForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Overtime record updated successfully.')
            return redirect('list_overtime_register')
    else:
        form = MinimumWagesOvertimeRegisterForm(instance=record)
        form.fields['attendance'].queryset = attendance.objects.filter(companyid=company)

    return render(request, 'Aapp/generic/form.html', {
        'form': form, 'company': company,
        'page_title': 'Edit Overtime Record',
        'cancel_url': reverse('list_overtime_register'),
    })


@login_required
def delete_overtime_register(request, ot_register_id):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    record = get_object_or_404(MinimumWagesOvertimeRegister, ot_register_id=ot_register_id,
                                attendance__companyid=company)

    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Overtime record deleted successfully.')
        return redirect('list_overtime_register')

    return render(request, 'Aapp/generic/confirm.html', {
        'company': company,
        'page_title': 'Delete Overtime Record',
        'confirm_message': f'Are you sure you want to delete the overtime record for '
                            f'<strong>{record.attendance.emp_code}</strong> on {record.ot_date}?',
        'cancel_url': reverse('list_overtime_register'),
    })
