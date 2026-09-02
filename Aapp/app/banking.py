"""
Aapp/app/banking.py
======================
NEFT/RTGS/IMPS bulk-upload file generation for salary disbursement, per
pt_upgrades.md: "include in banks", "NEFT/RTGS/IMPS file format for
banks", "Banking API (NOT NOW)".

Generates bulk payment files for a given salary_processing batch, in
both CSV and Excel (.xlsx), using each bank's own column layout. Direct
API disbursement is explicitly out of scope for this module (per spec)
— output is a file the Owner/HR uploads manually to their bank's
NetBanking/corporate portal.

Bank formats implemented (generic bulk-NEFT structure, common across
most Indian bank corporate portals — beneficiary account, IFSC, name,
amount, remarks are near-universal columns; exact column order/labels
vary slightly per bank and should be verified against the specific
bank's current template before first use in production):
  - HDFC Bank (ENet)
  - ICICI Bank (Corporate iCorp)
  - State Bank of India (CINB/INB)
  - Axis Bank (Corporate iConnect)
  - Generic (fallback for any other bank)

Auto/direct-to-bank-API disbursement is intentionally NOT built here —
matches "Banking API (NOT NOW)" in the spec.
"""

import csv
import io
from decimal import Decimal

from django.db import models
from django.forms import ModelForm
from django.forms.widgets import Select

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from Sapp.app.company import Company


BANK_FORMAT_CHOICES = [
    ('hdfc', 'HDFC Bank (ENet)'),
    ('icici', 'ICICI Bank (Corporate iCorp)'),
    ('sbi', 'State Bank of India (CINB)'),
    ('axis', 'Axis Bank (Corporate iConnect)'),
    ('generic', 'Generic NEFT/RTGS/IMPS'),
]

PAYMENT_MODE_CHOICES = [('NEFT', 'NEFT'), ('RTGS', 'RTGS'), ('IMPS', 'IMPS')]

# Column headers per bank format. All formats share the same underlying
# data (beneficiary account, IFSC, name, amount, remarks) — only the
# header labels and column order differ to match each bank's template.
BANK_COLUMN_LAYOUTS = {
    'hdfc': [
        'Beneficiary Account Number', 'Beneficiary IFSC', 'Beneficiary Name',
        'Amount', 'Payment Mode', 'Narration', 'Debit Account Number',
    ],
    'icici': [
        'Debit Account Number', 'Beneficiary Account Number', 'Receiver IFSC Code',
        'Beneficiary Customer Name', 'Transaction Amount', 'Transaction Currency',
        'Transaction Remarks', 'Payment Mode',
    ],
    'sbi': [
        'Sr No', 'Beneficiary Name', 'Beneficiary Account No', 'IFSC Code',
        'Amount', 'Mode', 'Remarks',
    ],
    'axis': [
        'Beneficiary Name', 'Account Number', 'IFSC Code', 'Amount',
        'Payment Type', 'Remarks', 'Debit Account',
    ],
    'generic': [
        'Beneficiary Name', 'Beneficiary Account Number', 'IFSC Code',
        'Amount', 'Payment Mode', 'Remarks',
    ],
}


class BankPaymentBatch(models.Model):
    """
    One row per generated bulk payment file — records what was
    generated, for audit/reference. The actual file bytes are not
    stored in the DB; only the metadata and row count are, to keep
    this table light. Regenerate the file on demand from the same
    salary_processing batch if needed again.
    """
    batch_id = models.AutoField(primary_key=True)
    company = models.ForeignKey(Company, on_delete=models.CASCADE, related_name='bank_payment_batches')
    salary_processing = models.ForeignKey('Aapp.salary_processing', on_delete=models.CASCADE,
                                           related_name='bank_payment_batches')
    bank_format = models.CharField(max_length=10, choices=BANK_FORMAT_CHOICES)
    payment_mode = models.CharField(max_length=4, choices=PAYMENT_MODE_CHOICES, default='NEFT')
    employee_count = models.PositiveIntegerField(default=0)
    total_amount = models.DecimalField(max_digits=14, decimal_places=2, default=0)
    generated_at = models.DateTimeField(auto_now_add=True)
    generated_by = models.CharField(max_length=50, null=True, blank=True)

    class Meta:
        app_label = 'Aapp'
        db_table = 'aa_bank_payment_batch'
        ordering = ['-generated_at']
        verbose_name = "Bank Payment Batch"

    def __str__(self):
        return f"Batch #{self.batch_id} - {self.get_bank_format_display()} - {self.employee_count} employees"


class BankPaymentBatchForm(ModelForm):
    class Meta:
        model = BankPaymentBatch
        fields = ['bank_format', 'payment_mode']
        widgets = {
            'bank_format': Select(attrs={'class': 'form-control'}),
            'payment_mode': Select(attrs={'class': 'form-control'}),
        }


def _get_payable_rows(processing_batch):
    """
    Returns list of dicts for every employee in this processing batch
    with net_pay > 0 and complete bank details on file. Employees
    missing account/IFSC are excluded and reported separately — a
    payment file with blank/invalid rows would be rejected by the bank
    portal outright.
    """
    from Aapp.app.salary_processing import salary_slip

    slips = (salary_slip.objects
             .filter(processing_id=processing_batch)
             .select_related('employee_id')
             .order_by('employee_id__employeecode'))

    rows, skipped = [], []
    for slip in slips:
        emp = slip.employee_id
        if slip.net_pay <= 0:
            continue
        if not emp.bank_account or not emp.bank_ifsc:
            skipped.append({'employee_code': emp.employeecode, 'name': emp.name, 'reason': 'Missing bank account/IFSC'})
            continue
        rows.append({
            'name': emp.name,
            'account': emp.bank_account,
            'ifsc': emp.bank_ifsc,
            'amount': slip.net_pay,
        })
    return rows, skipped


def generate_bank_csv(processing_batch, bank_format, payment_mode, company):
    """Returns (csv_bytes, row_count, total_amount, skipped_list)."""
    rows, skipped = _get_payable_rows(processing_batch)
    columns = BANK_COLUMN_LAYOUTS.get(bank_format, BANK_COLUMN_LAYOUTS['generic'])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)

    total = Decimal('0')
    for i, row in enumerate(rows, start=1):
        total += row['amount']
        writer.writerow(_format_row_for_bank(bank_format, i, row, payment_mode, company))

    return output.getvalue().encode('utf-8'), len(rows), total, skipped


def generate_bank_xlsx(processing_batch, bank_format, payment_mode, company):
    """Returns (xlsx_bytes, row_count, total_amount, skipped_list)."""
    rows, skipped = _get_payable_rows(processing_batch)
    columns = BANK_COLUMN_LAYOUTS.get(bank_format, BANK_COLUMN_LAYOUTS['generic'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment File"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    total = Decimal('0')
    for i, row in enumerate(rows, start=1):
        total += row['amount']
        values = _format_row_for_bank(bank_format, i, row, payment_mode, company)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=i + 1, column=col_idx, value=value)

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows), total, skipped


def _format_row_for_bank(bank_format, seq_no, row, payment_mode, company):
    """Returns a list of cell values in the exact column order for the given bank_format."""
    debit_account = company.account or ''

    if bank_format == 'hdfc':
        return [row['account'], row['ifsc'], row['name'], float(row['amount']),
                payment_mode, f"Salary - {row['name']}", debit_account]
    elif bank_format == 'icici':
        return [debit_account, row['account'], row['ifsc'], row['name'],
                float(row['amount']), 'INR', f"Salary Payment", payment_mode]
    elif bank_format == 'sbi':
        return [seq_no, row['name'], row['account'], row['ifsc'],
                float(row['amount']), payment_mode, f"Salary - {row['name']}"]
    elif bank_format == 'axis':
        return [row['name'], row['account'], row['ifsc'], float(row['amount']),
                payment_mode, f"Salary Payment", debit_account]
    else:  # generic
        return [row['name'], row['account'], row['ifsc'], float(row['amount']),
                payment_mode, f"Salary - {row['name']}"]


# =====================================================================
# VIEWS
# =====================================================================

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse


def _company(request):
    cid = request.session.get('selected_company_id')
    return Company.objects.filter(company_id=cid).first() if cid else None


@login_required
def list_bank_batches(request):
    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batches = BankPaymentBatch.objects.filter(company=company).select_related('salary_processing').order_by('-generated_at')
    rows = [{
        'cells': [
            b.batch_id, b.get_bank_format_display(), b.get_payment_mode_display(),
            b.employee_count, f"Rs. {b.total_amount}", b.generated_at.strftime('%d-%m-%Y %H:%M'),
        ],
        'actions': [
            {'url': reverse('download_bank_csv', args=[b.batch_id]), 'label': 'CSV'},
            {'url': reverse('download_bank_xlsx', args=[b.batch_id]), 'label': 'Excel'},
        ],
    } for b in batches]

    return render(request, 'Aapp/generic/list.html', {
        'page_title': 'Bank Payment Batches',
        'columns': ['Batch ID', 'Bank Format', 'Mode', 'Employees', 'Total Amount', 'Generated At'],
        'rows': rows, 'company': company,
        'add_url': reverse('select_processing_for_bank_file'), 'add_label': 'Generate New Batch',
        'empty_message': 'No payment batches generated yet.',
    })


@login_required
def select_processing_for_bank_file(request):
    """Step 1: pick which processed salary batch to generate the bank file from."""
    from Aapp.app.salary_processing import salary_processing

    company = _company(request)
    if not company:
        messages.warning(request, 'Please select a company first.')
        return redirect('aapp_dashboard')

    batches = salary_processing.objects.filter(company=company, status='processed').order_by('-year', '-month')
    return render(request, 'Aapp/works/select_processing_batch.html', {'batches': batches, 'company': company})


@login_required
def create_bank_batch(request, processing_id):
    """Step 2: pick bank format + mode, generate the file record."""
    from Aapp.app.salary_processing import salary_processing

    company = _company(request)
    processing_batch = get_object_or_404(salary_processing, processing_id=processing_id, company=company)

    if request.method == 'POST':
        form = BankPaymentBatchForm(request.POST)
        if form.is_valid():
            rows, skipped = _get_payable_rows(processing_batch)
            if not rows:
                messages.error(request, 'No employees with valid bank details found in this batch.')
                return redirect('select_processing_for_bank_file')

            total = sum((r['amount'] for r in rows), start=Decimal('0'))
            batch = form.save(commit=False)
            batch.company = company
            batch.salary_processing = processing_batch
            batch.employee_count = len(rows)
            batch.total_amount = total
            batch.generated_by = request.user.username
            batch.save()

            if skipped:
                messages.warning(
                    request,
                    f"{len(skipped)} employee(s) skipped due to missing bank details: " +
                    ", ".join(s['employee_code'] for s in skipped)
                )
            messages.success(request, f'Bank payment batch generated: {len(rows)} employees, Rs. {total}.')
            return redirect('list_bank_batches')
    else:
        form = BankPaymentBatchForm()

    return render(request, 'Aapp/works/create_bank_batch.html', {
        'form': form, 'processing_batch': processing_batch, 'company': company
    })


@login_required
def download_bank_csv(request, batch_id):
    company = _company(request)
    batch = get_object_or_404(BankPaymentBatch, batch_id=batch_id, company=company)
    csv_bytes, _, _, _ = generate_bank_csv(batch.salary_processing, batch.bank_format, batch.payment_mode, company)
    return HttpResponse(csv_bytes, content_type='text/csv', headers={
        'Content-Disposition': f'attachment; filename="bank_payment_{batch.bank_format}_{batch_id}.csv"'
    })


@login_required
def download_bank_xlsx(request, batch_id):
    company = _company(request)
    batch = get_object_or_404(BankPaymentBatch, batch_id=batch_id, company=company)
    xlsx_bytes, _, _, _ = generate_bank_xlsx(batch.salary_processing, batch.bank_format, batch.payment_mode, company)
    return HttpResponse(xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         headers={'Content-Disposition': f'attachment; filename="bank_payment_{batch.bank_format}_{batch_id}.xlsx"'})
