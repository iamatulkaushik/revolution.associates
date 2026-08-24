"""
Cxapp/app/banking.py
=======================
NEFT/RTGS/IMPS bulk-upload file generation for the Cxapp portal —
mirrors Aapp.app.banking, scoped to CxOwnerProfile/CxSalary/CxEmployee.

Debit account: CxOwnerProfile.company.account/ifsc (Sapp.Company fields
already used by Aapp's banking module — no new fields needed here,
since a Cxapp company IS a Sapp.Company underneath).

Beneficiary account: CxEmployeeBanking.account_number — this is an
EncryptedCharField (AES-256-GCM); Django's ORM decrypts it transparently
on attribute access, so this module reads emp.banking.account_number
exactly like any plaintext field. It is never written back to any
other table or logged — only placed directly into the generated
CSV/XLSX bytes returned to the browser for download.

Bank column layouts are identical to Aapp's (HDFC/ICICI/SBI/Axis/
generic) — imported directly rather than duplicated, since column
layout is a pure data structure with no Cxapp/Aapp-specific coupling.
"""

import csv
import io
from decimal import Decimal

from django.db import models
from django.forms import ModelForm
from django.forms.widgets import Select

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from Aapp.app.banking import BANK_FORMAT_CHOICES, PAYMENT_MODE_CHOICES, BANK_COLUMN_LAYOUTS


class CxBankPaymentBatch(models.Model):
    """Audit record of a generated bulk payment file — file bytes not stored, regenerated on demand."""
    batch_id = models.AutoField(primary_key=True)
    company = models.ForeignKey('Cxapp.CxOwnerProfile', on_delete=models.CASCADE, related_name='bank_payment_batches')
    cx_salary_batch_month = models.PositiveSmallIntegerField()
    cx_salary_batch_year = models.PositiveIntegerField()
    bank_format = models.CharField(max_length=10, choices=BANK_FORMAT_CHOICES)
    payment_mode = models.CharField(max_length=4, choices=PAYMENT_MODE_CHOICES, default='NEFT')
    employee_count = models.PositiveIntegerField(default=0)
    total_amount = models.DecimalField(max_digits=14, decimal_places=2, default=0)
    generated_at = models.DateTimeField(auto_now_add=True)
    generated_by = models.CharField(max_length=50, blank=True)

    class Meta:
        app_label = 'Cxapp'
        db_table = 'cx_bank_payment_batch'
        ordering = ['-generated_at']
        verbose_name = "Bank Payment Batch"

    def __str__(self):
        return f"Batch #{self.batch_id} - {self.get_bank_format_display()} - {self.employee_count} employees"


class CxBankPaymentBatchForm(ModelForm):
    class Meta:
        model = CxBankPaymentBatch
        fields = ['bank_format', 'payment_mode']
        widgets = {
            'bank_format': Select(attrs={'class': 'form-control'}),
            'payment_mode': Select(attrs={'class': 'form-control'}),
        }


def _get_payable_rows(company_owner_profile, month, year):
    """
    Returns (rows, skipped) for every CxSalary in the given month/year
    with total_amount > 0 and complete bank details on file (banking
    record exists + bank_ifsc set). Employees missing either are
    skipped and reported separately.
    """
    from Cxapp.app.process import CxSalary

    salaries = (CxSalary.objects
                .filter(attendance__company=company_owner_profile,
                        salary_month=month, salary_year=year)
                .select_related('employee'))

    rows, skipped = [], []
    for salary in salaries:
        emp = salary.employee
        if salary.total_amount <= 0:
            continue

        banking = getattr(emp, 'banking', None)
        if not banking or not banking.bank_ifsc or not banking.account_number:
            skipped.append({'employee_code': emp.employee_code, 'name': emp.name,
                             'reason': 'Missing bank account/IFSC'})
            continue

        rows.append({
            'name': emp.name,
            'account': banking.account_number,  # transparently decrypted by EncryptedCharField
            'ifsc': banking.bank_ifsc,
            'amount': salary.total_amount,
        })
    return rows, skipped


def _format_row_for_bank(bank_format, seq_no, row, payment_mode, debit_account):
    if bank_format == 'hdfc':
        return [row['account'], row['ifsc'], row['name'], float(row['amount']),
                payment_mode, f"Salary - {row['name']}", debit_account]
    elif bank_format == 'icici':
        return [debit_account, row['account'], row['ifsc'], row['name'],
                float(row['amount']), 'INR', "Salary Payment", payment_mode]
    elif bank_format == 'sbi':
        return [seq_no, row['name'], row['account'], row['ifsc'],
                float(row['amount']), payment_mode, f"Salary - {row['name']}"]
    elif bank_format == 'axis':
        return [row['name'], row['account'], row['ifsc'], float(row['amount']),
                payment_mode, "Salary Payment", debit_account]
    else:
        return [row['name'], row['account'], row['ifsc'], float(row['amount']),
                payment_mode, f"Salary - {row['name']}"]


def generate_bank_csv(company_owner_profile, month, year, bank_format, payment_mode):
    rows, skipped = _get_payable_rows(company_owner_profile, month, year)
    columns = BANK_COLUMN_LAYOUTS.get(bank_format, BANK_COLUMN_LAYOUTS['generic'])
    debit_account = company_owner_profile.company.account or ''

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)

    total = Decimal('0')
    for i, row in enumerate(rows, start=1):
        total += row['amount']
        writer.writerow(_format_row_for_bank(bank_format, i, row, payment_mode, debit_account))

    return output.getvalue().encode('utf-8'), len(rows), total, skipped


def generate_bank_xlsx(company_owner_profile, month, year, bank_format, payment_mode):
    rows, skipped = _get_payable_rows(company_owner_profile, month, year)
    columns = BANK_COLUMN_LAYOUTS.get(bank_format, BANK_COLUMN_LAYOUTS['generic'])
    debit_account = company_owner_profile.company.account or ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment File"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    total = Decimal('0')
    for i, row in enumerate(rows, start=1):
        total += row['amount']
        values = _format_row_for_bank(bank_format, i, row, payment_mode, debit_account)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=i + 1, column=col_idx, value=value)

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows), total, skipped


# =====================================================================
# VIEWS
# =====================================================================

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse


def _can_manage_payroll(request):
    if getattr(request, 'cx_sub_user', None) is None:
        return True
    return request.cx_sub_user.get_role_permissions().get('wages', False)


def cxapp_list_bank_batches(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_list_bank_batches)(request)


def _list_bank_batches(request):
    if not _can_manage_payroll(request):
        messages.error(request, 'You do not have permission to view bank batches.')
        return redirect('cxapp_dashboard')

    owner_profile = request.cx_owner_profile
    batches = CxBankPaymentBatch.objects.filter(company=owner_profile).order_by('-generated_at')
    return render(request, 'Cxapp/banking/batch_list.html', {'batches': batches})


def cxapp_select_salary_for_bank_file(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_select_salary_for_bank_file)(request)


def _select_salary_for_bank_file(request):
    from Cxapp.app.process import CxSalary

    if not _can_manage_payroll(request):
        messages.error(request, 'You do not have permission to generate bank files.')
        return redirect('cxapp_dashboard')

    owner_profile = request.cx_owner_profile
    batches = (CxSalary.objects
               .filter(attendance__company=owner_profile)
               .values('salary_month', 'salary_year')
               .distinct()
               .order_by('-salary_year', '-salary_month'))

    return render(request, 'Cxapp/banking/select_batch.html', {'batches': batches})


def cxapp_create_bank_batch(request, month, year):
    from Cxapp.views import cx_login_required
    return cx_login_required(_create_bank_batch)(request, month, year)


def _create_bank_batch(request, month, year):
    owner_profile = request.cx_owner_profile

    if request.method == 'POST':
        form = CxBankPaymentBatchForm(request.POST)
        if form.is_valid():
            rows, skipped = _get_payable_rows(owner_profile, month, year)
            if not rows:
                messages.error(request, 'No employees with valid bank details found for this month.')
                return redirect('cxapp_select_salary_for_bank_file')

            total = sum((r['amount'] for r in rows), start=Decimal('0'))
            batch = form.save(commit=False)
            batch.company = owner_profile
            batch.cx_salary_batch_month = month
            batch.cx_salary_batch_year = year
            batch.employee_count = len(rows)
            batch.total_amount = total
            batch.generated_by = getattr(request.cx_sub_user, 'username', 'Owner')
            batch.save()

            if skipped:
                messages.warning(
                    request,
                    f"{len(skipped)} employee(s) skipped due to missing bank details: " +
                    ", ".join(s['employee_code'] for s in skipped)
                )
            messages.success(request, f'Bank payment batch generated: {len(rows)} employees, Rs. {total}.')
            return redirect('cxapp_list_bank_batches')
    else:
        form = CxBankPaymentBatchForm()

    return render(request, 'Cxapp/banking/create_batch.html', {
        'form': form, 'month': month, 'year': year
    })


def cxapp_download_bank_csv(request, batch_id):
    from Cxapp.views import cx_login_required
    return cx_login_required(_download_bank_csv)(request, batch_id)


def _download_bank_csv(request, batch_id):
    owner_profile = request.cx_owner_profile
    batch = get_object_or_404(CxBankPaymentBatch, batch_id=batch_id, company=owner_profile)
    csv_bytes, _, _, _ = generate_bank_csv(
        owner_profile, batch.cx_salary_batch_month, batch.cx_salary_batch_year,
        batch.bank_format, batch.payment_mode
    )
    return HttpResponse(csv_bytes, content_type='text/csv', headers={
        'Content-Disposition': f'attachment; filename="bank_payment_{batch.bank_format}_{batch_id}.csv"'
    })


def cxapp_download_bank_xlsx(request, batch_id):
    from Cxapp.views import cx_login_required
    return cx_login_required(_download_bank_xlsx)(request, batch_id)


def _download_bank_xlsx(request, batch_id):
    owner_profile = request.cx_owner_profile
    batch = get_object_or_404(CxBankPaymentBatch, batch_id=batch_id, company=owner_profile)
    xlsx_bytes, _, _, _ = generate_bank_xlsx(
        owner_profile, batch.cx_salary_batch_month, batch.cx_salary_batch_year,
        batch.bank_format, batch.payment_mode
    )
    return HttpResponse(xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         headers={'Content-Disposition': f'attachment; filename="bank_payment_{batch.bank_format}_{batch_id}.xlsx"'})
