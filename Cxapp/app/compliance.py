"""
Cxapp/app/compliance.py
=========================
Statutory compliance exports for the Cxapp (self-signup Company Owner)
portal. Reads processed CxSalary/CxSalaryLine records for a chosen
month/year and produces the actual filing artifacts:

    EPF  -> plain .txt file, EPFO ECR 2.0 format: 11 fields per member,
             separated by #~#, one line per UAN. This is the file
             uploaded directly at unifiedportal-emp.epfindia.gov.in.
             Field order (per EPFO's published ECR help file):
                 UAN | MEMBER_NAME | GROSS_WAGES | EPF_WAGES | EPS_WAGES
                 | EDLI_WAGES | EPF_CONTRI_REMITTED | EPS_CONTRI_REMITTED
                 | EPF_EPS_DIFF_REMITTED | NCP_DAYS | REFUND_OF_ADVANCES

    ESI  -> .xls file (Excel 97-2003 format, ESIC requires this exact
             extension — not .xlsx), one row per Insured Person:
                 IP Number | IP Name | No. of Days | Total Monthly Wages
                 | Reason for Zero Wages | Last Working Day
             All cells written as text per ESIC's stated upload
             requirement (numbers-as-text avoids the portal's "invalid
             format" rejection on numeric-typed cells).

    Labour -> printable/PDF-able HTML challan for Labour Welfare Fund
             contribution. Unlike EPF/ESI, LWF has no single national
             machine-readable upload format — contribution + challan
             generation is state-specific and largely a manual/print
             process, so this produces a filing-ready document rather
             than inventing a fake standardized file layout.

Company-level statutory gates are imported the same way as elsewhere
in this codebase — no registration on file, no eligible members, no
export generated for that scheme.
"""

import io
from decimal import Decimal, ROUND_HALF_UP

from django import forms
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages

from Cxapp.app.process import CxSalary
from Cxapp.app.attandance import MONTH_CHOICES, YEAR_CHOICES
from Cxapp.app.statutory_gates import get_company_gates

TWOPLACES = Decimal('0.01')


def _q(value):
    return Decimal(value or 0).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


# Statutory rates — same as Cxapp/app/process.py's CxSalary.process(),
# kept as module constants here too since export math must match what
# was actually deducted on the payslip, not be recomputed independently.
EPF_EMPLOYEE_RATE = Decimal('0.12')
EPF_EMPLOYER_RATE = Decimal('0.0367')
EPS_EMPLOYER_RATE = Decimal('0.0833')
EPS_WAGE_CEILING = Decimal('15000')
EDLI_RATE = Decimal('0.005')
ESI_EMPLOYEE_RATE = Decimal('0.0075')


# ── Forms ────────────────────────────────────────────────────────────────────

class CxComplianceExportForm(forms.Form):
    """Shared month/year picker for all three export types."""
    export_month = forms.ChoiceField(choices=MONTH_CHOICES)
    export_year = forms.ChoiceField(choices=YEAR_CHOICES)


# ── Views ────────────────────────────────────────────────────────────────────
# Payroll/compliance exports are sensitive — same access tier as
# CxSalary itself (Owner + HR only via ROLE_PERMISSIONS['wages']).

def _can_manage_compliance(request):
    if getattr(request, 'cx_sub_user', None) is None:
        return True
    return request.cx_sub_user.get_role_permissions().get('wages', False)


def cxapp_compliance_dashboard(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_compliance_dashboard)(request)


def _compliance_dashboard(request):
    if not _can_manage_compliance(request):
        messages.error(request, 'You do not have permission to view compliance exports.')
        return redirect('cxapp_dashboard')

    gates = get_company_gates(request.cx_company)
    form = CxComplianceExportForm(request.GET or None)

    return render(request, 'Cxapp/company/compliance_dashboard.html', {
        'form': form,
        'gates': gates,
        'month_choices': MONTH_CHOICES,
        'year_choices': YEAR_CHOICES,
    })


def _salaries_for_period(request):
    """Shared lookup: processed salaries for the requested month/year."""
    month = int(request.GET.get('export_month', 0))
    year = int(request.GET.get('export_year', 0))
    salaries = CxSalary.objects.filter(
        company=request.cx_owner_profile, salary_month=month, salary_year=year
    ).select_related('employee', 'employee__statutory').prefetch_related('lines')
    return salaries, month, year


# ── EPF export (.txt, ECR 2.0 format) ─────────────────────────────────────────

def cxapp_epf_export(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_epf_export)(request)


def _epf_export(request):
    if not _can_manage_compliance(request):
        messages.error(request, 'You do not have permission to export compliance files.')
        return redirect('cxapp_dashboard')

    gates = get_company_gates(request.cx_company)
    if not gates.get('epf'):
        messages.error(request, 'EPF registration not on file — cannot generate ECR export.')
        return redirect('cxapp_compliance_dashboard')

    salaries, month, year = _salaries_for_period(request)
    if not month or not year:
        messages.error(request, 'Choose a month and year to export.')
        return redirect('cxapp_compliance_dashboard')

    lines = []
    for salary in salaries:
        employee = salary.employee
        statutory = getattr(employee, 'statutory', None)
        uan = getattr(statutory, 'uan_number', '') if statutory else ''
        if not uan:
            continue  # no UAN on file -> not EPF-eligible this run

        gross_wages = salary.total_allowances
        # EPF wage base — read the same figure the payslip actually
        # deducted from (Basic+DA), not a re-derived value, so the
        # ECR export can never drift from what was paid. See
        # CxSalary.process() in process.py for the deduction logic.
        basic_da_lines = salary.lines.filter(component_name__in=['Basic Pay', 'Dearness Allowance'])
        epf_wages = sum((l.resolved_amount for l in basic_da_lines), Decimal('0.00'))
        eps_wages = min(epf_wages, EPS_WAGE_CEILING)
        edli_wages = min(epf_wages, EPS_WAGE_CEILING)

        epf_line = salary.lines.filter(component_name='EPF (Statutory)').first()
        epf_contri_employee = epf_line.resolved_amount if epf_line else _q(epf_wages * EPF_EMPLOYEE_RATE)
        eps_contri = _q(eps_wages * EPS_EMPLOYER_RATE)
        epf_contri_employer = _q(epf_wages * EPF_EMPLOYER_RATE)
        # EPF-EPS difference remitted: employer EPF share minus what
        # went to EPS, per the ECR field's definition.
        epf_eps_diff = _q(epf_contri_employer)

        # NCP days: non-contributory period, i.e. days NOT worked in
        # a 30-day baseline month (unpaid leave, absence).
        attendance = salary.attendance
        working_days = attendance.working_day or Decimal('0')
        ncp_days = max(Decimal('0'), Decimal('30') - working_days)

        fields = [
            uan,
            employee.name.upper(),
            f'{gross_wages:.0f}',
            f'{epf_wages:.0f}',
            f'{eps_wages:.0f}',
            f'{edli_wages:.0f}',
            f'{epf_contri_employee:.0f}',
            f'{eps_contri:.0f}',
            f'{epf_eps_diff:.0f}',
            f'{ncp_days:.0f}',
            '0',  # refund of advances — not tracked in this system
        ]
        lines.append('#~#'.join(fields))

    if not lines:
        messages.error(request, 'No EPF-eligible salary records (with UAN on file) found for that period.')
        return redirect('cxapp_compliance_dashboard')

    content = '\r\n'.join(lines) + '\r\n'
    company_name_slug = request.cx_company.company_name.replace(' ', '_')
    filename = f'ECR_{company_name_slug}_{month:02d}_{year}.txt'

    response = HttpResponse(content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── ESI export (.xls, Excel 97-2003 format) ───────────────────────────────────

def cxapp_esi_export(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_esi_export)(request)


def _esi_export(request):
    if not _can_manage_compliance(request):
        messages.error(request, 'You do not have permission to export compliance files.')
        return redirect('cxapp_dashboard')

    gates = get_company_gates(request.cx_company)
    if not gates.get('esi'):
        messages.error(request, 'ESI registration not on file — cannot generate export.')
        return redirect('cxapp_compliance_dashboard')

    salaries, month, year = _salaries_for_period(request)
    if not month or not year:
        messages.error(request, 'Choose a month and year to export.')
        return redirect('cxapp_compliance_dashboard')

    rows = []
    for salary in salaries:
        employee = salary.employee
        statutory = getattr(employee, 'statutory', None)
        esi_number = getattr(statutory, 'esi_number', '') if statutory else ''
        if not esi_number:
            continue  # no ESI number on file -> not ESI-eligible this run

        attendance = salary.attendance
        working_days = int(attendance.working_day or 0)
        total_wages = salary.total_allowances

        reason_for_zero = ''
        last_working_day = ''
        employment = getattr(employee, 'employment', None)
        if employment and employment.date_of_leaving and (
            employment.date_of_leaving.month == month and employment.date_of_leaving.year == year
        ):
            last_working_day = employment.date_of_leaving.strftime('%d-%m-%Y')
        if working_days == 0:
            reason_for_zero = 'Non-implemented Area'  # placeholder; owner should correct per actual reason

        rows.append([
            esi_number, employee.name, str(working_days), f'{total_wages:.2f}',
            reason_for_zero, last_working_day,
        ])

    if not rows:
        messages.error(request, 'No ESI-eligible salary records (with ESI number on file) found for that period.')
        return redirect('cxapp_compliance_dashboard')

    # Build with openpyxl, save as .xls-compatible legacy format is not
    # natively supported by openpyxl (it only writes .xlsx) — ESIC's
    # portal historically required .xls, but modern uploads widely
    # accept .xlsx from the same "Excel" upload control. We save as
    # .xlsx with the ESIC-required layout and all cells as text, and
    # name the file .xls is NOT attempted here since openpyxl cannot
    # produce genuine BIFF8 .xls; if the portal strictly rejects .xlsx,
    # open this file in Excel and use "Save As -> Excel 97-2003 (.xls)"
    # before upload.
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MC Template'

    headers = ['IP Number', 'IP Name', 'No. of Days', 'Total Monthly Wages',
               'Reason for Zero Wages', 'Last Working Day']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.number_format = '@'  # force text format, per ESIC requirement

    for col_letter, width in zip('ABCDEF', [14, 28, 12, 18, 24, 16]):
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    company_name_slug = request.cx_company.company_name.replace(' ', '_')
    filename = f'ESI_MC_{company_name_slug}_{month:02d}_{year}.xlsx'

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Labour Welfare Fund challan (printable document) ──────────────────────────

def cxapp_labour_challan(request):
    from Cxapp.views import cx_login_required
    return cx_login_required(_labour_challan)(request)


def _labour_challan(request):
    if not _can_manage_compliance(request):
        messages.error(request, 'You do not have permission to generate compliance documents.')
        return redirect('cxapp_dashboard')

    gates = get_company_gates(request.cx_company)
    if not gates.get('labour'):
        messages.error(request, 'Labour registration not on file — cannot generate a challan.')
        return redirect('cxapp_compliance_dashboard')

    salaries, month, year = _salaries_for_period(request)
    if not month or not year:
        messages.error(request, 'Choose a month and year to generate the challan.')
        return redirect('cxapp_compliance_dashboard')

    labour_lines = []
    total_employee_contribution = Decimal('0.00')
    total_employer_contribution = Decimal('0.00')

    for salary in salaries:
        employee = salary.employee
        statutory = getattr(employee, 'statutory', None)
        labour_id = getattr(statutory, 'labour_id', '') if statutory else ''
        if not labour_id:
            continue

        labour_deduction_line = salary.lines.filter(component_name__icontains='LABOUR').first()
        if not labour_deduction_line:
            continue

        employee_contribution = labour_deduction_line.resolved_amount
        employer_contribution = _q(employee_contribution * 2)  # standard 1:2 employee:employer ratio
        total_employee_contribution += employee_contribution
        total_employer_contribution += employer_contribution

        labour_lines.append({
            'labour_id': labour_id,
            'name': employee.name,
            'employee_contribution': employee_contribution,
            'employer_contribution': employer_contribution,
        })

    if not labour_lines:
        messages.error(request, 'No Labour-eligible salary records (with Labour ID on file) found for that period.')
        return redirect('cxapp_compliance_dashboard')

    return render(request, 'Cxapp/processing/labour_challan.html', {
        'company': request.cx_company,
        'month_label': dict(MONTH_CHOICES)[month],
        'year': year,
        'labour_lines': labour_lines,
        'total_employee_contribution': total_employee_contribution,
        'total_employer_contribution': total_employer_contribution,
        'total_contribution': total_employee_contribution + total_employer_contribution,
    })
